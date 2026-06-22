from django.shortcuts import render
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from .models import (
    Country,
    University,
    Course,
    TeamMember,
    Testimonial,
    Blog,
    Service,
    Category,
    GalleryImage,
    ContactMessage,
    Application,
    CourseCategory,
    Inquiry,
)
from django.core.paginator import Paginator
from .forms import (
    CountryForm,
    UniversityForm,
    CourseForm,
    TeamMemberForm,
    TestimonialForm,
    BlogForm,
    ServiceForm,
    CategoryForm,
    GalleryImageForm,
    ContactMessageForm,
    ApplicationForm,
    CourseCategoryForm,
    InquiryForm,
)
from django.contrib import messages
import math
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import random
import os
from django.urls import reverse
import datetime
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
import csv
from django.db.models import Count
from django.db.models.functions import TruncMonth
import datetime
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
import json
from .forms import NewsletterForm
from django.utils.dateparse import parse_date
from .models import NewsletterSubscriber
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
import pandas as pd
from .models import ChatbotUser
from django.db.models.functions import Lower
from student_management.decorators import admin_login_required


# User Side >>>>>>>>>>>>>>>>>>>>>>>>>


def index(request):
    course_categories = CourseCategory.objects.all()
    countries = Country.objects.all()

    countriess = Country.objects.order_by('?')[:5]
    team_members = TeamMember.objects.all()[:6]
    universities = University.objects.all()
    
    services = Service.objects.all()

    footer_service = Service.objects.all()[:4]

    # services = Service.objects.all()[:5]
    testimonials = Testimonial.objects.all()
    blogs = Blog.objects.all().order_by("-created_at")[:2]
    print(">>>>>>>>>>>>>>>>>>>>.Countries in index:", countries)
    print(universities)
    return render(
        request,
        "index.html",
        {
            "countries": countries,
            "team_members": team_members,
            "countriess": countriess,
            "services": services,
            "universities": universities,
            "testimonials": testimonials,
            "blogs": blogs,
            "course_categories": course_categories,
            "footer_service": footer_service,
        },
    )




def country_details(request, slug):
    course_categories = CourseCategory.objects.all()
    services = Service.objects.all()
    footer_service = Service.objects.all()[:4]
    
    country = get_object_or_404(Country, slug=slug)
    countries = Country.objects.all()
    other_countries = Country.objects.all()
    
    # Get all universities for this country
    universities_list = country.universities.all()
    
    # Pagination - 4 universities per page
    paginator = Paginator(universities_list, 4)
    page = request.GET.get("page", 1)
    
    try:
        universities_page = paginator.page(page)
    except PageNotAnInteger:
        universities_page = paginator.page(1)
    except EmptyPage:
        universities_page = paginator.page(paginator.num_pages)
    
    # AJAX request - return JSON with university data
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        universities_data = []
        for uni in universities_page:
            universities_data.append({
                'name': uni.name,
                'slug': uni.slug,
                'image_url': uni.image.url if uni.image else '',
                'country_name': uni.country.name,
                'description': uni.description[:150] + '...' if len(uni.description) > 150 else uni.description,
            })
        
        return JsonResponse({
            'universities': universities_data,
            'has_previous': universities_page.has_previous(),
            'has_next': universities_page.has_next(),
            'current_page': universities_page.number,
            'total_pages': paginator.num_pages,
            'start_index': universities_page.start_index(),
            'end_index': universities_page.end_index(),
            'total_count': paginator.count,
            'previous_page': universities_page.previous_page_number() if universities_page.has_previous() else None,
            'next_page': universities_page.next_page_number() if universities_page.has_next() else None,
        })
    
    context = {
        "country": country,
        "countries": countries,
        "universities": universities_page,
        "course_categories": course_categories,
        "services": services,
        "footer_service": footer_service,
        "other_countries": other_countries,
    }
    return render(request, "country-details.html", context)




def university_detail(request, slug):
    course_categories = CourseCategory.objects.all()
    countries = Country.objects.all()
    universities = University.objects.all()

    services = Service.objects.all()
    # university = get_object_or_404(University, id=pk)
    university = get_object_or_404(University, slug=slug)
    courses = Course.objects.filter(university=university)
    # application_form = ApplicationForm.objects.last()

    footer_service = Service.objects.all()[:4]

    # Calculate course groups (4 courses per slide)
    paginator = Paginator(courses, 4)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "university": university,
        "courses": courses,
        "page_obj": page_obj,
        "course_categories": course_categories,
        "services": services,
        "countries": countries,
        "universities": universities,
        "footer_service": footer_service,
    }
    return render(request, "university_detail.html", context)


def about(request):
    course_categories = CourseCategory.objects.all()
    countries = Country.objects.all()
    universities = University.objects.all()[:6]
    services = Service.objects.all()
    blogs = Blog.objects.all()[:6]

    team_members = TeamMember.objects.all()[:6]
    testimonials = Testimonial.objects.all()

    footer_service = Service.objects.all()[:4]

    context = {
        "universities": universities,
        "course_categories": course_categories,
        "services": services,
        "blogs": blogs,
        "countries": countries,
        "team_members": team_members,
        "testimonials": testimonials,
        "footer_service": footer_service,
    }

    return render(request, "about.html", context)


def blog_details(request, slug):

    course_categories = CourseCategory.objects.all()
    countries = Country.objects.all()
    universities = University.objects.all()
    services = Service.objects.all()

    footer_service = Service.objects.all()[:4]

    # Get the current blog post
    # blog = get_object_or_404(Blog, id=blog_id)
    blog = get_object_or_404(Blog, slug=slug)

    # Get recent blogs for sidebar (excluding current blog)
    recent_blogs = Blog.objects.exclude(id=blog.id).order_by("-created_at")[:5]

    # optional >>>>>>>..
    related_blogs = Blog.objects.exclude(id=blog.id).order_by("-created_at")[:3]

    context = {
        "blog": blog,
        "recent_blogs": recent_blogs,
        "related_blogs": related_blogs,
        "universities": universities,
        "course_categories": course_categories,
        "services": services,
        "countries": countries,
        "footer_service": footer_service,
    }

    return render(request, "blog_details.html", context)


def contact_submit(request):
    if request.method == "POST":
        form = ContactMessageForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Your message has been sent successfully!")
            return redirect(request.META.get("HTTP_REFERER", "/"))
        else:
            messages.error(
                request, "There was an error sending your message. Please try again."
            )
    else:
        form = ContactMessageForm()
    return render(request, "university_detail.html", {"form": form})


def service_detail(request, slug):
    # blog = blog.objects.all()
    services = Service.objects.all()
    other_services = Service.objects.all()
    course_categories = CourseCategory.objects.all()
    countries = Country.objects.all()
    # service = get_object_or_404(Service, pk=pk)
    service = get_object_or_404(Service, slug=slug)

    footer_service = Service.objects.all()[:4]

    context = {
        # 'blog': blog,
        "course_categories": course_categories,
        "service": service,
        "countries": countries,
        "services": services,
        "footer_service": footer_service,
        "other_services": other_services,
    }

    return render(request, "service_detail.html", context)


def gallery(request):
    blog = Blog.objects.all()
    course_categories = CourseCategory.objects.all()
    countries = Country.objects.all()
    services = Service.objects.all()

    categories = Category.objects.all()

    footer_service = Service.objects.all()[:4]

    # Pagination for All images
    page_number_all = request.GET.get("page", 1)
    all_images = GalleryImage.objects.all()
    paginator_all = Paginator(all_images, 9)  # 6 per page
    images_all = paginator_all.get_page(page_number_all)

    # Pagination for category images
    category_images_list = []
    for category in categories:
        cat_page = request.GET.get(f"page_{category.id}", 1)
        images = category.images.all()
        paginator = Paginator(images, 6)
        category_images_list.append((category, paginator.get_page(cat_page)))

    context = {
        "categories": categories,
        "images_all": images_all,
        "category_images_list": category_images_list,
        "blog": blog,
        "course_categories": course_categories,
        "services": services,
        "countries": countries,
        "footer_service": footer_service,
    }

    return render(request, "gallery.html", context)


def inquiry_view(request):
    if request.method == "POST":
        form = InquiryForm(request.POST)
        if form.is_valid():
            # Save the form to database
            form.save()
            print(form, "this is my form>>>>>>>>>>>>>>>>>")
            messages.success(request, "Your inquiry has been submitted successfully!")

            # Redirect to prevent form resubmission
            return redirect("contact_us")
        else:
            # Form is invalid, show errors
            messages.error(request, "Please correct the errors below.")
    else:
        form = InquiryForm()

    course_categories = CourseCategory.objects.all()
    services = Service.objects.all()
    countries = Country.objects.all()
    footer_service = Service.objects.all()[:4]

    context = {
        "form": form,
        "course_categories": course_categories,
        "countries": countries,
        "services": services,
        "footer_service":footer_service
    }

    return render(request, "contact-us.html", context)


def subscribe_newsletter(request):
    if request.method == "POST":
        form = NewsletterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Thank you for subscribing to our newsletter!")
            return redirect(request.META.get("HTTP_REFERER", "/"))
        else:
            messages.error(request, "Please enter a valid email address.")
            return redirect(request.META.get("HTTP_REFERER", "/"))


def apply_form(request):
    countries = Country.objects.all()
    courses = Course.objects.all()

    course_categories = CourseCategory.objects.all()
    services = Service.objects.all()
    footer_service = Service.objects.all()[:4]

    if request.method == "POST":
        form = ApplicationForm(request.POST)  # include FILES
        if form.is_valid():
            form.save()
            messages.success(
                request, "Your application has been submitted successfully!"
            )
            return redirect("apply_form")
        else:
            messages.error(request, " Please correct the errors below.")
    else:
        form = ApplicationForm()

    context = {
        "form": form,
        "countries": countries,
        "courses": courses,
        "course_categories": course_categories,
        "services": services,
        "footer_service": footer_service,
        
    }
    return render(request, "apply-form.html", context)


def course_category_detail(request, slug):
    course_categories = CourseCategory.objects.all()
    other_course_categories = CourseCategory.objects.all()
    # category = get_object_or_404(CourseCategory, id=category_id)
    category = get_object_or_404(CourseCategory, slug=slug)
    courses = Course.objects.filter(category=category).order_by("-id")
    countries = Country.objects.all()
    services = Service.objects.all()

    footer_service = Service.objects.all()[:4]

    # Paginate courses
    paginator = Paginator(courses, 4)  # 4 courses per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "category": category,
        "page_obj": page_obj,
        "courses": courses,
        "course_categories": course_categories,
        "countries": countries,
        "services": services,
        "footer_service": footer_service,
        "other_course_categories": other_course_categories,
    }
    return render(request, "category_detials.html", context)


def index_blog(request):
    blogs = Blog.objects.all().order_by("-created_at")

    course_categories = CourseCategory.objects.all()
    countries = Country.objects.all()

    services = Service.objects.all()

    footer_service = Service.objects.all()[:4]
    paginator = Paginator(blogs, 4)  # 2 blogs per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "course_categories": course_categories,
        "page_obj": page_obj,
        "countries": countries,
        "services": services,
        "footer_service": footer_service,
    }

    return render(request, "blogs.html", context)


# admin side >>>>>>>>>>>>>>>>>>>>>>>>>>>>>


@login_required(login_url="admin_login")
def subscriber_list(request):

    subscribers = NewsletterSubscriber.objects.all().order_by("-subscribed_at")

    paginator = Paginator(subscribers, 7)  # 7 per page
    page_number = request.GET.get("page")
    subscribers_page = paginator.get_page(page_number)

    return render(
        request,
        "admin_pages/subscriber_list.html",
        {
            "subscribers": subscribers_page,
        },
    )


@admin_login_required
def delete_subscriber(request, pk):
    subscriber = get_object_or_404(NewsletterSubscriber, pk=pk)
    if request.method == "POST":
        subscriber.delete()
        messages.success(request, "Subscriber deleted successfully.")
        return redirect("subscriber_list")
    return redirect("subscriber_list")


# Admin-side list messages (custom page)
@admin_login_required
def admin_contacts(request):
    contacts = ContactMessage.objects.all().order_by("-created_at")

    # Date filter
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    if start_date and end_date:
        contacts = contacts.filter(created_at__date__range=[start_date, end_date])

    paginator = Paginator(contacts, 10)  # 10 per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "admin_pages/contact_list.html", {"contacts": page_obj})


@admin_login_required
def delete_contact(request, contact_id):
    contact = get_object_or_404(ContactMessage, id=contact_id)
    if request.method == "POST":
        contact.delete()
        messages.success(request, "Contact message deleted successfully.")
        return redirect("admin_contacts")
    messages.error(request, "Invalid request.")
    return redirect("admin_contacts")


# Export to Excel/CSV
@admin_login_required
def export_contacts_excel(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    contacts = ContactMessage.objects.all()
    if start_date and end_date:
        contacts = contacts.filter(created_at__date__range=[start_date, end_date])

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Header row
    ws.append(["Name", "Email", "Phone", "Message", "Created At"])

    # Data rows
    for c in contacts:
        ws.append(
            [
                c.name,
                c.email,
                c.phone,
                c.message,
                c.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="contacts_{datetime.date.today()}.xlsx"'
    )

    wb.save(response)
    return response


@admin_login_required
def admin_dashboard(request):
    # Stats
    stats = {
        "applications_count": Application.objects.count(),
        "inquiries_count": Inquiry.objects.count(),
        "countries_count": Country.objects.count(),
    }

    contacts_count = ContactMessage.objects.count()

    applications_data = (
        Application.objects.annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )

    # Testimonials per month
    testimonials_data = (
        Testimonial.objects.annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )

    applications_labels = [x["month"].strftime("%b %Y") for x in applications_data]
    applications_counts = [x["count"] for x in applications_data]

    testimonials_labels = [x["month"].strftime("%b %Y") for x in testimonials_data]
    testimonials_counts = [x["count"] for x in testimonials_data]

    # Recent Data
    recent_applications = Application.objects.select_related("course").order_by(
        "-created_at"
    )[:5]
    recent_inquiries = Inquiry.objects.order_by("-created_at")[:5]

    # Top Countries (by applications count)
    top_countries = Country.objects.annotate(app_count=Count("application")).order_by(
        "-app_count"
    )[:5]

    context = {
        "stats": stats,
        "contacts_count": contacts_count,
        "recent_applications": recent_applications,
        "recent_inquiries": recent_inquiries,
        "top_countries": top_countries,
        "applications_labels": applications_labels,
        "applications_counts": applications_counts,
        "testimonials_labels": testimonials_labels,
        "testimonials_counts": testimonials_counts,
    }
    return render(request, "admin_pages/admin-dashboard.html", context)


def page_404(request, exception):
    return render(request, "404.html", status=404)


@admin_login_required
def countries_list(request):
    countries = Country.objects.all().order_by('order', 'id')
    
    context = {
        'countries': countries,
    }
    return render(request, 'admin_pages/country_list.html', context)

@admin_login_required
def country_create(request):
    if request.method == "POST":
        form = CountryForm(request.POST, request.FILES)
        if form.is_valid():
            name = form.cleaned_data["name"].strip()

            if Country.objects.filter(name__iexact=name).exists():
                messages.error(request, f"Country '{name}' already exists.")
            else:
                form.save()
                messages.success(request, "Country created successfully.")
                return redirect("country_list")
        else:
            messages.error(request, "fix error below the images!")
    else:
        form = CountryForm()
    return render(request, "admin_pages/create-country.html", {"form": form})


@admin_login_required
def country_update(request, pk):
    country = get_object_or_404(Country, pk=pk)
    if request.method == "POST":
        form = CountryForm(request.POST, request.FILES, instance=country)
        if form.is_valid():
            country = form.save()

            messages.success(request, "Country updated successfully.")
            return redirect("country_list")
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        form = CountryForm(instance=country)
    return render(
        request, "admin_pages/country_list.html", {"form": form, "country": country}
    )


@admin_login_required
def country_delete(request, pk):
    country = get_object_or_404(Country, pk=pk)
    if request.method == "POST":
        country.delete()
        messages.success(request, "Country deleted successfully.")
        return redirect("country_list")
    return render(request, "admin_pages/country_list.html", {"country": country})


@admin_login_required
def add_university(request):
    countries = Country.objects.all()

    if request.method == "POST":
        form = UniversityForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "University added successfully!")
            return redirect("uni-list")
    else:
        form = UniversityForm()

    context = {"form": form, "countries": countries}
    return render(request, "admin_pages/create_university.html", context)


@admin_login_required
def universities_list(request):
    universities = University.objects.select_related('country').order_by('order', 'id')
    countries = Country.objects.all()
    
    context = {
        'universities': universities,
        'countries': countries,
    }
    return render(request, 'admin_pages/university_list.html', context)


@admin_login_required
# Update university
def update_university(request, pk):
    university = get_object_or_404(University, pk=pk)
    if request.method == "POST":
        form = UniversityForm(request.POST, request.FILES, instance=university)
        if form.is_valid():
            form.save()
            messages.success(request, "University updated successfully!")
            return redirect("uni-list")
    else:
        form = UniversityForm(instance=university)

    context = {"form": form, "university": university}
    return render(request, "admin_pages/update_university.html", context)


@admin_login_required
#  Delete university
def delete_university(request, pk):
    university = get_object_or_404(University, pk=pk)
    if request.method == "POST":  # confirmation before delete
        university.delete()
        messages.success(request, "University deleted successfully!")
        return redirect("uni-list")

    context = {"university": university}
    return render(request, "admin_pages/delete_university.html", context)



@admin_login_required
def course_list(request):
    courses = Course.objects.select_related('university', 'category', 'university__country').order_by('order', 'id')
    categories = CourseCategory.objects.all()
    universities = University.objects.all()
    
    context = {
        'courses': courses,
        'categories': categories,
        'universities': universities,
    }
    return render(request, 'admin_pages/course_list.html', context)


@admin_login_required
# Add new course
def course_add(request):

    categories = CourseCategory.objects.all().order_by("-created_at")
    print("ebtererrere")
    if request.method == "POST":
        form = CourseForm(request.POST, request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Course added successfully!")
            return redirect("course_list")  # update with your course list url name

    else:
        form = CourseForm()

    universities = University.objects.select_related("country").all()

    return render(
        request,
        "admin_pages/course_create.html",
        {"form": form, "universities": universities, "categories": categories},
    )


@admin_login_required
#  Update course
def course_update(request, pk):
    course = get_object_or_404(Course, pk=pk)
    universities = University.objects.select_related(
        "country"
    ).all()  # Get all universities
    categories = CourseCategory.objects.all().order_by("-created_at")
    print(categories)

    if request.method == "POST":
        form = CourseForm(request.POST, request.FILES, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, "Course updated successfully!")
            return redirect("course_list")
    else:
        form = CourseForm(instance=course)

    # In case of form errors, we need to render the course_list template with the form and the universities
    courses = Course.objects.select_related("university__country").all()
    return render(
        request,
        "admin_pages/course_list.html",
        {
            "form": form,
            "course": course,
            "courses": courses,
            "universities": universities,  # pass to template
            "categories": categories,
        },
    )


@admin_login_required
def course_delete(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == "POST":
        course.delete()
        messages.success(request, "Course deleted successfully!")
        return redirect("course_list")
    return render(request, "admin_pages/course_list.html", {"course": course})


@admin_login_required
def list_team(request):
    """Display all team members with pagination"""
    team_members_list = TeamMember.objects.all().order_by(Lower("name"))
    paginator = Paginator(team_members_list, 6)  # Show 10 per page
    page_number = request.GET.get("page")
    team_members = paginator.get_page(
        page_number
    )  # handles invalid pages automatically

    context = {
        "team_members": team_members,
        "title": "Team Members",
    }
    return render(request, "admin_pages/team_list.html", context)


@admin_login_required
def create_team(request):
    if request.method == "POST":
        form = TeamMemberForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Team member added successfully!")
            return redirect("team_list")
        messages.error(request, "Please correct the errors below.")
    else:
        form = TeamMemberForm()

    return render(request, "admin_pages/add_team.html", {"form": form})


@admin_login_required
def edit_team_member(request, pk):
    team_member = get_object_or_404(TeamMember, pk=pk)
    if request.method == "POST":
        form = TeamMemberForm(request.POST, request.FILES, instance=team_member)
        if form.is_valid():
            form.save()
            messages.success(request, "Team member updated successfully!")
            return redirect("team_list")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = TeamMemberForm(instance=team_member)

    return render(
        request,
        "admin_pages/team_list.html",
        {"form": form, "team_member": team_member},
    )


@admin_login_required
def delete_team_member(request, pk):
    team_member = get_object_or_404(TeamMember, pk=pk)
    if request.method == "POST":
        team_member.delete()
        messages.success(request, "Team member deleted successfully!")
        return redirect("team_list")

    return render(request, "admin_pages/team_list.html", {"team_member": team_member})


@admin_login_required
def testimonial_list(request):
    testimonials_list = Testimonial.objects.all().order_by(Lower("name"))
    paginator = Paginator(testimonials_list, 6)
    page_number = request.GET.get("page")
    testimonials = paginator.get_page(page_number)

    return render(
        request, "admin_pages/review_list.html", {"testimonials": testimonials}
    )


@admin_login_required
def testimonial_create(request):
    if request.method == "POST":
        form = TestimonialForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Testimonial added successfully!")
            return redirect("testimonial_list")
    else:
        form = TestimonialForm()
    return render(request, "admin_pages/create_review.html", {"form": form})


@admin_login_required
def testimonial_update(request, pk):
    testimonial = get_object_or_404(Testimonial, pk=pk)
    if request.method == "POST":
        form = TestimonialForm(request.POST, request.FILES, instance=testimonial)
        if form.is_valid():
            form.save()
            messages.success(request, "Testimonial updated successfully!")
            return redirect("testimonial_list")
    else:
        form = TestimonialForm(instance=testimonial)
    return render(
        request,
        "admin_pages/review_list.html",
        {"form": form, "testimonial": testimonial},
    )

@admin_login_required
def testimonial_delete(request, pk):
    testimonial = get_object_or_404(Testimonial, pk=pk)
    if request.method == "POST":
        testimonial.delete()
        messages.success(request, "Testimonial deleted successfully!")
        return redirect("testimonial_list")
    return render(request, "admin_pages/review_list.html", {"testimonial": testimonial})


# --------- Services ---------
@admin_login_required
def service_list(request):
    services = Service.objects.all().order_by('order', 'id')
    return render(request, "admin_pages/service_list.html", {"services": services})


@admin_login_required
def service_create(request):
    if request.method == "POST":
        form = ServiceForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Service added successfully!")
            return redirect("service_list")
    else:
        form = ServiceForm()
    return render(request, "admin_pages/create_service.html", {"form": form})


@admin_login_required
def service_update(request, pk):
    service = get_object_or_404(Service, pk=pk)
    if request.method == "POST":
        form = ServiceForm(request.POST, request.FILES, instance=service)
        if form.is_valid():
            form.save()
            messages.success(request, "Service updated successfully!")
            return redirect("service_list")
    else:
        form = ServiceForm(instance=service)
    return render(
        request, "admin_pages/service_list.html", {"form": form, "service": service}
    )


@admin_login_required
def service_delete(request, pk):
    service = get_object_or_404(Service, pk=pk)
    if request.method == "POST":
        service.delete()
        messages.success(request, "Service deleted successfully!")
        return redirect("service_list")

    # If it's a GET request, show confirmation (handled by the modal)
    return redirect("service_list")


# --------- Blogs ---------
@admin_login_required
def blog_list(request):
    blogs_qs = Blog.objects.all().order_by("title")# newest first

    paginator = Paginator(blogs_qs, 6)
    page_number = request.GET.get("page")
    blogs = paginator.get_page(page_number)  # gives a Page object

    return render(request, "admin_pages/blog_list.html", {"blogs": blogs})


@admin_login_required
def blog_create(request):
    if request.method == "POST":
        form = BlogForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Blog added successfully!")
            return redirect("blog_list")
    else:
        form = BlogForm()
    return render(request, "admin_pages/create_blog.html", {"form": form})


@admin_login_required
def blog_update(request, pk):
    blog = get_object_or_404(Blog, pk=pk)
    if request.method == "POST":
        form = BlogForm(request.POST, request.FILES, instance=blog)
        if form.is_valid():
            form.save()
            messages.success(request, "Blog updated successfully!")
            return redirect("blog_list")
    else:
        form = BlogForm(instance=blog)
    return render(request, "admin_pages/create_blog.html", {"form": form, "blog": blog})


@admin_login_required
def blog_delete(request, pk):
    blog = get_object_or_404(Blog, pk=pk)
    if request.method == "POST":
        blog.delete()
        messages.success(request, "Blog deleted successfully!")
        return redirect("blog_list")
    return render(request, "admin_pages/create_blog.html", {"blog": blog})


@admin_login_required
def gallery_images(request):
    categories = Category.objects.all().prefetch_related("images")

    category_pages = {}
    for category in categories:
        images_qs = category.images.all().order_by("-uploaded_at")
        paginator = Paginator(images_qs, 8)  # 8 images per page
        page_number = request.GET.get(f"page_{category.id}", 1)

        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        category_pages[category.id] = page_obj

    return render(
        request,
        "admin_pages/image_list.html",
        {
            "categories": categories,
            "category_pages": category_pages,
        },
    )


@admin_login_required
def add_image(request):
    categories = Category.objects.all()

    if request.method == "POST":
        category_id = request.POST.get("category")
        category = Category.objects.get(id=category_id)
        files = request.FILES.getlist("images")
        print("FILES:", request.FILES)  # Should show uploaded files
        print("FILES count:", len(request.FILES.getlist("images")))

        for file in files:
            GalleryImage.objects.create(
                category=category,
                title=file.name,  # default title = filename
                image=file,
            )
        messages.success(request, "Images uploaded succesfully")
        return redirect("list_image")
    return render(request, "admin_pages/add_image.html", {"categories": categories})


@admin_login_required
def application_list(request):
    apps = Application.objects.all().order_by("-created_at")

    # Date filter
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    if start_date and end_date:
        apps = apps.filter(created_at__date__range=[start_date, end_date])

    # Pagination
    paginator = Paginator(apps, 10)  # 10 per page
    page = request.GET.get("page")
    applications = paginator.get_page(page)

    return render(
        request,
        "admin_pages/application_list.html",
        {
            "applications": applications,
        },
    )

@admin_login_required
def delete_application(request, app_id):
    app = get_object_or_404(Application, id=app_id)
    if request.method == "POST":
        app.delete()
        messages.success(request, "Application deleted successfully.")
        return redirect("application_list")
    messages.error(request, "Invalid request.")
    return redirect("application_list")


def get_courses_by_country(request, country_id):
    courses = Course.objects.filter(university__country_id=country_id).values(
        "id", "title"
    )
    return JsonResponse(list(courses), safe=False)


@admin_login_required
def course_category_list_create(request):
    categories = CourseCategory.objects.all().order_by("-created_at")

    if request.method == "POST":
        form = CourseCategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, " Category added successfully!")
            return redirect("course_category_list")
    else:
        form = CourseCategoryForm()

    context = {
        "categories": categories,
        "form": form,
    }
    return render(request, "admin_pages/course_category.html", context)


@admin_login_required
def export_applications_excel(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    applications = Application.objects.all()
    if start_date and end_date:
        applications = applications.filter(
            created_at__date__range=[start_date, end_date]
        )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Header row
    ws.append(
        [
            "Full Name",
            "Email",
            "Phone",
            "Qualification",
            "Marks",
            "Course",
            "Country",
            "Created At",
        ]
    )

    # Data rows
    for app in applications:
        ws.append(
            [
                app.full_name,
                app.email,
                app.phone,
                app.qualification or "",
                float(app.marks) if app.marks else "",
                app.course.title if app.course else "",
                app.country.name if app.country else "",
                app.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="applications_{datetime.date.today()}.xlsx"'
    )

    wb.save(response)
    return response


@admin_login_required
def course_category_list(request):
    categories = CourseCategory.objects.all().order_by('order', 'id')
    
    context = {
        'categories': categories,
    }
    return render(request, 'admin_pages/course_category_list.html', context)


@admin_login_required
def course_category_update(request, pk):
    category = get_object_or_404(CourseCategory, pk=pk)
    if request.method == "POST":
        form = CourseCategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Category updated successfully!")
            return redirect("course_category_list")
    else:
        form = CourseCategoryForm(instance=category)
    return render(
        request,
        "admin_pages/course_category_update.html",
        {"form": form, "category": category},
    )


@admin_login_required
def course_category_delete(request, pk):
    category = get_object_or_404(CourseCategory, pk=pk)
    if request.method == "POST":
        category.delete()
        messages.success(request, "Category deleted successfully!")
        return redirect("course_category_list")
    return render(
        request, "admin_pages/course_category_delete.html", {"category": category}
    )


@admin_login_required
def delete_image(request, image_id):
    image = get_object_or_404(GalleryImage, id=image_id)

    if request.method == "POST":
        image.delete()
        messages.success(request, "Image deleted successfully")
        return redirect("list_image")

    return render(request, "admin_pages/image_list.html", {"image": image})


@admin_login_required
def category_list(request):
    categories = Category.objects.all().order_by("-created_at")
    paginator = Paginator(categories, 10)
    page_number = request.GET.get("page")
    categories = paginator.get_page(page_number)
    return render(request, "admin_pages/category_list.html", {"categories": categories})


@admin_login_required
def add_category(request):
    if request.method == "POST":
        name = request.POST.get("name")
        if name:
            Category.objects.create(name=name)
            return redirect("category_list")
    return render(request, "admin_pages/add_category.html")


@admin_login_required
def update_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        category.name = request.POST.get("name")
        category.save()
        return redirect("category_list")
    return redirect("category_list")


@admin_login_required
def delete_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        category.delete()
        return redirect("category_list")
    return redirect("category_list")


@admin_login_required
def inquiry_list(request):
    inquiries = Inquiry.objects.all().order_by("-created_at")

    # Date filtering
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date:
        inquiries = inquiries.filter(created_at__date__gte=start_date)
    if end_date:
        inquiries = inquiries.filter(created_at__date__lte=end_date)

    # Pagination
    paginator = Paginator(inquiries, 7)  # Show 20 inquiries per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "inquiries": page_obj,
    }
    return render(request, "admin_pages/inquiry_list.html", context)


@admin_login_required
def delete_inquiry(request, pk):
    inquiry = get_object_or_404(Inquiry, pk=pk)
    if request.method == "POST":
        inquiry.delete()
        return redirect("inquiry_list")  # change to your listing view name
    return redirect("inquiry_list")


@admin_login_required
def export_inquiries_excel(request):
    # Get filter parameters
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # Filter inquiries based on date range
    inquiries = Inquiry.objects.all()

    if start_date:
        inquiries = inquiries.filter(created_at__date__gte=start_date)
    if end_date:
        inquiries = inquiries.filter(created_at__date__lte=end_date)

    # Prepare data for Excel
    data = []
    for inquiry in inquiries:
        data.append(
            {
                "Name": inquiry.name,
                "Email": inquiry.email,
                "Phone": inquiry.phone if inquiry.phone else "-",
                "Message": inquiry.message,
                "Date": inquiry.created_at.strftime("%Y-%m-%d %H:%M"),
            }
        )

    # Create DataFrame
    df = pd.DataFrame(data)

    # Create HTTP response with Excel file
    response = HttpResponse(content_type="application/ms-excel")
    filename = f"contact_inquiries_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Write DataFrame to Excel
    df.to_excel(response, index=False, sheet_name="Contact Inquiries")

    return response


def admin_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        if not username or not password:
            messages.error(request, "Both fields are required.")
            return render(request, "authenticate/login.html")

        user = authenticate(request, username=username, password=password)

        if user is not None and user.is_staff:  # only staff users
            login(request, user)
            messages.success(request, f"Welcome {user.username}!")
            print("login sucesfull")
            return redirect("student_management:dashboard")  # change this to your dashboard URL
        else:
            messages.error(request, "Invalid credentials or not an admin.")

    return render(request, "authenticate/login.html")


@admin_login_required
def admin_logout(request):
    logout(request)
    messages.success(request, "You have been logged out.")
    return redirect("admin_login")


@csrf_exempt
@require_http_methods(["POST"])
def save_chatbot_data(request):
    try:
        data = json.loads(request.body)

        user_phone = data.get("phone")

        preference = data.get("preference", "")

        user = None
        if user_phone:
            user = ChatbotUser.objects.filter(phone=user_phone).first()

        if user:
            user.name = data.get("name", user.name)
            user.phone = data.get("phone", user.phone)
            user.qualification = data.get("qualification", user.qualification)
            user.year_completed = data.get("year_completed", user.year_completed)
            user.preference = preference or ""
            user.save()
        else:
            user = ChatbotUser.objects.create(
                name=data.get("name"),
                phone=data.get("phone"),
                qualification=data.get("qualification"),
                year_completed=data.get("year_completed"),
                preference=preference or "",
            )

        return JsonResponse(
            {
                "status": "success",
                "user_id": user.id,
                "message": "User data saved successfully",
            }
        )

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)})


@require_http_methods(["GET"])
def get_chatbot_options(request):

    try:
        # Get services
        services = list(Service.objects.all().values("id", "title", "description")[:10])

        # Get countries
        countries = list(Country.objects.all().values("id", "name", "description")[:10])

        # Get universities with country names
        all_universities = list(
            University.objects.all().values(
                "id", "name", "country__name", "description"
            )[:10]
        )

        # Get ccategory with courses
        course_categories = list(
            CourseCategory.objects.all().values("id", "name", "description")[:10]
        )

        universities = random.sample(all_universities, min(6, len(all_universities)))

        transformed_categories = []
        for category in course_categories:
            # Get courses for this category
            category_courses = Course.objects.filter(category_id=category["id"]).values(
                "id", "title", "university__name", "duration"
            )[:10]

            transformed_categories.append(
                {
                    "id": category["id"],
                    "title": category["name"],  # Map 'name' to 'title'
                    "description": category["description"],
                    "courses": list(category_courses),  # Include courses list
                }
            )

        quick_menu_options = [
            {"title": " Services", "type": "services", "category": "main"},
            {"title": " Course Categories", "type": "courses", "category": "main"},
            {"title": " Universities", "type": "universities", "category": "main"},
            {"title": " Countries", "type": "countries", "category": "main"},
        ]

        return JsonResponse(
            {
                "quick_menu": quick_menu_options,
                "services": services,
                "countries": countries,
                "universities": universities,
                "courses": transformed_categories,
            }
        )
    except Exception as e:
        print(f"Error in get_chatbot_options: {str(e)}")
        return JsonResponse(
            {
                "quick_menu": [],
                "services": [],
                "countries": [],
                "universities": [],
                "courses": [],
            }
        )


@admin_login_required
def chatbot_enquiries_list(request):
    enquiries = ChatbotUser.objects.all().order_by("-created_at")

    # Date filtering
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date:
        enquiries = enquiries.filter(created_at__date__gte=start_date)
    if end_date:
        enquiries = enquiries.filter(created_at__date__lte=end_date)

    # Pagination
    paginator = Paginator(enquiries, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "chatbot_enquiries": page_obj,
    }
    return render(request, "admin_pages/chatuser_list.html", context)


@admin_login_required
def delete_chatbot_enquiry(request, enquiry_id):
    enquiry = get_object_or_404(ChatbotUser, id=enquiry_id)
    if request.method == "POST":
        enquiry.delete()
        return redirect("chatbot_enquiries_list")
    return redirect("chatbot_enquiries_list")


import openpyxl


@admin_login_required
def export_chatbot_enquiries_excel(request):
    enquiries = ChatbotUser.objects.all().order_by("-created_at")

    # Date filtering
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date:
        enquiries = enquiries.filter(created_at__date__gte=start_date)
    if end_date:
        enquiries = enquiries.filter(created_at__date__lte=end_date)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chatbot Enquiries"

    # Headers
    headers = ["Name", "Email", "Phone", "Qualification", "Year Completed", "Date"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Data
    for row, enquiry in enumerate(enquiries, 2):
        ws.cell(row=row, column=1, value=enquiry.name)
        ws.cell(row=row, column=2, value=enquiry.email)
        ws.cell(row=row, column=3, value=enquiry.phone or "")
        ws.cell(row=row, column=4, value=enquiry.qualification or "")
        ws.cell(row=row, column=5, value=enquiry.year_completed or "")
        ws.cell(row=row, column=7, value=enquiry.created_at.strftime("%Y-%m-%d %H:%M"))

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=chatbot_enquiries.xlsx"
    wb.save(response)

    return response




from django.views.decorators.http import require_POST

@admin_login_required
@require_POST
def reorder_universities(request):
    """Handle AJAX request to update university order"""
    try:
        data = json.loads(request.body)
        order_data = data.get('order', [])
        
        # Update each university's order
        for index, uni_id in enumerate(order_data):
            University.objects.filter(id=uni_id).update(order=index)
        
        return JsonResponse({
            'status': 'success',
            'message': 'Universities reordered successfully'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)
        
        
@csrf_exempt
@require_POST
def reorder_services(request):
    try:
        data = json.loads(request.body)
        order = data.get('order', [])
        
        # Update the order of services
        for index, service_id in enumerate(order):
            service = Service.objects.get(id=service_id)
            service.order = index  # Assuming you have an 'order' field
            service.save()
            
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
    
# In your views.py
@csrf_exempt
@require_POST
@admin_login_required
def reorder_countries(request):
    try:
        data = json.loads(request.body)
        order = data.get('order', [])
        
        # Update the order of countries
        for index, country_id in enumerate(order):
            country = Country.objects.get(id=country_id)
            country.order = index
            country.save()
            
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
    
@csrf_exempt
@require_POST
@admin_login_required
def reorder_course_categories(request):
    try:
        data = json.loads(request.body)
        order = data.get('order', [])
        
        # Update the order of categories
        for index, category_id in enumerate(order):
            category = CourseCategory.objects.get(id=category_id)
            category.order = index
            category.save()
            
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})
    
    
@csrf_exempt
@require_POST
@admin_login_required
def reorder_courses(request):
    try:
        data = json.loads(request.body)
        order = data.get('order', [])
        
        # Update the order of courses
        for index, course_id in enumerate(order):
            course = Course.objects.get(id=course_id)
            course.order = index
            course.save()
            
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})