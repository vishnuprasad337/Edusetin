from django.shortcuts import render, get_object_or_404, redirect
from django.http import FileResponse, Http404, JsonResponse
from django.conf import settings
from django.db.models import Q,Avg
import os
import io
import re
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from student_portal.models import Student
from .models import Subject, Question, QuestionMedia, MediaLibrary, PendingMediaReference,SubModule, Subject
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from student_portal.models import ExamAttempt,QuizAttempt
from django.core.paginator import Paginator
from django.db.models import Q
from .decorators import admin_login_required





# ─────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────

def normalize_question(text):
    text = text.lower().strip()
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(
        r'\s*([?.!,;:+\-*/=])\s*',
        r'\1',
        text
    )
    return text


# Maps Question requires_* field names to QuestionMedia media_type values
SLOT_MAP = [
    ('requires_question_image', 'QUESTION'),
    ('requires_option_a_image', 'OPTION_A'),
    ('requires_option_b_image', 'OPTION_B'),
    ('requires_option_c_image', 'OPTION_C'),
    ('requires_option_d_image', 'OPTION_D'),
    ('requires_option_e_image', 'OPTION_E'),
]

# Human-readable labels for each slot
SLOT_LABELS = {
    'QUESTION': 'Question Image',
    'OPTION_A': 'Option A Image',
    'OPTION_B': 'Option B Image',
    'OPTION_C': 'Option C Image',
    'OPTION_D': 'Option D Image',
    'OPTION_E': 'Option E Image',
}

ALLOWED_IMAGE_EXTENSIONS = {'jpg', 'jpeg', 'png', 'webp'}


def _validate_image_extension(filename):
    """Returns True if the file extension is allowed."""
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    return ext in ALLOWED_IMAGE_EXTENSIONS


def _recalculate_media_status(question):
    """
    Recalculate and save media_uploaded for a question.
    Returns (required_count, uploaded_count, missing_types).
    """
    required_types = {
        media_type
        for field, media_type in SLOT_MAP
        if getattr(question, field)
    }
    uploaded_types = set(
        question.media_files.values_list('media_type', flat=True)
    )
    missing_types = required_types - uploaded_types

    if required_types:
        question.media_uploaded = not bool(missing_types)
    else:
        question.media_uploaded = False

    question.save(update_fields=['media_uploaded'])
    return len(required_types), len(uploaded_types & required_types), missing_types


def _parse_bool_cell(value):
    """Parse YES/TRUE/1 (case-insensitive) as True, everything else as False."""
    if value is None:
        return False
    try:
        if pd.isna(value):
            return False
    except (TypeError, ValueError):
        pass
    return str(value).strip().upper() in ('YES', 'TRUE', '1')


def _get_media_slot_status(question, media_map):
    """
    Returns a list of dicts for each slot:
      {field, media_type, label, required, uploaded, media, pending_ref}
    """
    # Build a map of pending refs keyed by media_type
    pending_refs = {
        ref.media_type: ref
        for ref in question.pending_media_refs.all()
    }

    slots = []
    for field, media_type in SLOT_MAP:
        required = getattr(question, field)
        uploaded = media_type in media_map
        slots.append({
            'field': field,
            'media_type': media_type,
            'label': SLOT_LABELS[media_type],
            'required': required,
            'uploaded': uploaded,
            'media': media_map.get(media_type),
            'pending_ref': pending_refs.get(media_type),
        })
    return slots


# ─────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────

from django.db.models import Sum, Count
from django.utils import timezone

@admin_login_required
def student_management_dashboard(request):
    now = timezone.now()

    # ── Existing stats ────────────────────────────────────────────────
    total_students    = Student.objects.count()
    total_subjects    = Subject.objects.count()
    total_questions   = Question.objects.count()
    total_pyqs        = Question.objects.filter(source='PYQ').count()
    active_questions  = Question.objects.filter(is_active=True).count()
    inactive_questions = Question.objects.filter(is_active=False).count()

    recent_students  = Student.objects.select_related('user').order_by('-created_at')[:5]
    recent_subjects  = Subject.objects.order_by('-created_at')[:5]
    recent_questions = Question.objects.select_related('subject').order_by('-created_at')[:5]

    # ── Subscription plans ────────────────────────────────────────────
    active_plans       = SubscriptionPlan.objects.filter(is_active=True)
    total_active_plans = active_plans.count()
    recent_plans       = active_plans.order_by('-created_at')[:5]

    # ── Payment stats ─────────────────────────────────────────────────
    total_payments    = Payment.objects.count()
    successful_payments = Payment.objects.filter(status=Payment.STATUS_SUCCESS)
    pending_payments  = Payment.objects.filter(status=Payment.STATUS_PENDING).count()
    failed_payments   = Payment.objects.filter(status=Payment.STATUS_FAILED).count()

    total_revenue     = successful_payments.aggregate(
        total=Sum('amount')
    )['total'] or 0

    # Active subscriptions — paid and not yet expired
    active_subscriptions = Payment.objects.filter(
        status=Payment.STATUS_SUCCESS,
        expires_at__gt=now,
    ).count()

    # Revenue this month
    monthly_revenue = Payment.objects.filter(
        status=Payment.STATUS_SUCCESS,
        paid_at__year=now.year,
        paid_at__month=now.month,
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Recent payments
    recent_payments = Payment.objects.select_related(
        'student__user', 'plan'
    ).order_by('-created_at')[:5]

    return render(request, 'student_management/dashboard.html', {
        # existing
        'total_students':       total_students,
        'total_subjects':       total_subjects,
        'total_questions':      total_questions,
        'total_pyqs':           total_pyqs,
        'active_questions':     active_questions,
        'inactive_questions':   inactive_questions,
        'recent_students':      recent_students,
        'recent_subjects':      recent_subjects,
        'recent_questions':     recent_questions,

        # plans
        'total_active_plans':   total_active_plans,
        'active_plans':         active_plans,
        'recent_plans':         recent_plans,

        # payments
        'total_payments':       total_payments,
        'pending_payments':     pending_payments,
        'failed_payments':      failed_payments,
        'total_revenue':        total_revenue,
        'monthly_revenue':      monthly_revenue,
        'active_subscriptions': active_subscriptions,
        'recent_payments':      recent_payments,
    })


# ─────────────────────────────────────────────
# STUDENTS
# ─────────────────────────────────────────────

@admin_login_required
def student_list(request):
    students = Student.objects.select_related('user').order_by('-created_at')
    return render(request, 'student_management/student_list.html', {'students': students})


@admin_login_required
def student_detail(request, id):
    student = get_object_or_404(Student.objects.select_related('user'), id=id)

    # Mock test (exam) attempts — only count completed/submitted ones
    exam_attempts = ExamAttempt.objects.filter(
        student=student,
        status=ExamAttempt.STATUS_SUBMITTED,
    )
    quiz_attempts = QuizAttempt.objects.filter(
        student=student,
        status=QuizAttempt.STATUS_SUBMITTED,
    )

    total_mock_tests = exam_attempts.count()
    total_quizzes = quiz_attempts.count()
    total_results = total_mock_tests + total_quizzes

    # Average score across both exam and quiz attempts
    exam_avg = exam_attempts.aggregate(avg=Avg('percentage'))['avg'] or 0
    quiz_avg = quiz_attempts.aggregate(avg=Avg('percentage'))['avg'] or 0
    combined_count = total_mock_tests + total_quizzes
    if combined_count:
        average_score = round(
            ((exam_avg * total_mock_tests) + (quiz_avg * total_quizzes)) / combined_count,
            1,
        )
    else:
        average_score = 0

    
    active_subscription = student.payments.filter(
        status=Payment.STATUS_SUCCESS,
        expires_at__gt=timezone.now(),
    ).select_related('plan').order_by('-expires_at').first()

    return render(request, 'student_management/student_detail.html', {
        'student': student,
        'total_mock_tests': total_mock_tests,
        'total_results': total_results,
        'average_score': average_score,
        'active_subscription': active_subscription,
    })
@admin_login_required
def student_activate(request, id):
    student = get_object_or_404(Student, id=id)
    if student.is_active:
        messages.warning(request, "Student is already active.")
    else:
        student.is_active = True
        student.save()
        student.user.is_active = True
        student.user.save()
        messages.success(request, "Student activated successfully.")
    return redirect('student_management:student_detail', id=student.id)


@admin_login_required
def student_deactivate(request, id):
    student = get_object_or_404(Student, id=id)
    if not student.is_active:
        messages.warning(request, "Student is already inactive.")
    else:
        student.is_active = False
        student.save()
        student.user.is_active = False
        student.user.save()
        messages.warning(request, "Student deactivated successfully.")
    return redirect('student_management:student_detail', id=student.id)


@admin_login_required
def student_delete(request, id):
    student = get_object_or_404(Student, id=id)
    if request.method == 'POST':
        user = student.user
        student.delete()
        user.delete()
        messages.success(request, "Student deleted successfully.")
        return redirect('student_management:student_list')
    return redirect('student_management:student_detail', id=student.id)


# ─────────────────────────────────────────────
# SUBJECTS
# ─────────────────────────────────────────────

@admin_login_required
def subject_list(request):
    subjects = Subject.objects.order_by('-created_at')
    return render(request, 'student_management/subject_list.html', {'subjects': subjects})

@admin_login_required
def subject_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        image = request.FILES.get('image')

        if not name:
            messages.error(request, "Subject name is required.")
            return render(request, 'student_management/subject_create.html', {
                'post_data': request.POST
            })

        if Subject.objects.filter(name__iexact=name).exists():
            messages.error(request, f"A subject with the name \"{name}\" already exists.")
            return render(request, 'student_management/subject_create.html', {
                'post_data': request.POST
            })

        subject = Subject.objects.create(
            name=name,
            description=description or None,
            is_active=is_active,
            image=image,
        )
        messages.success(request, f"Subject \"{subject.name}\" created successfully.")
        return redirect('student_management:subject_list')

    return render(request, 'student_management/subject_create.html', {})


@admin_login_required
def subject_edit(request, id):
    subject = get_object_or_404(Subject, id=id)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        image = request.FILES.get('image')

        if not name:
            messages.error(request, "Subject name is required.")
            return render(request, 'student_management/subject_edit.html', {'subject': subject})

        if Subject.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, f"Another subject with the name \"{name}\" already exists.")
            return render(request, 'student_management/subject_edit.html', {'subject': subject})

        subject.name = name
        subject.description = description or None
        subject.is_active = is_active
        if image:
            subject.image = image
        subject.save()
        messages.success(request, f"Subject \"{subject.name}\" updated successfully.")
        return redirect('student_management:subject_list')

    return render(request, 'student_management/subject_edit.html', {'subject': subject})


@admin_login_required
def subject_activate(request, id):
    subject = get_object_or_404(Subject, id=id)
    if subject.is_active:
        messages.warning(request, "Subject is already active.")
    else:
        subject.is_active = True
        subject.save()
        messages.success(request, "Subject activated successfully.")
    return redirect('student_management:subject_list')


@admin_login_required
def subject_deactivate(request, id):
    subject = get_object_or_404(Subject, id=id)
    if not subject.is_active:
        messages.warning(request, "Subject is already inactive.")
    else:
        subject.is_active = False
        subject.save()
        messages.warning(request, "Subject deactivated successfully.")
    return redirect('student_management:subject_list')


@admin_login_required
def subject_delete(request, id):
    subject = get_object_or_404(Subject, id=id)
    if request.method == 'POST':
        name = subject.name
        subject.delete()
        messages.success(request, f"Subject \"{name}\" deleted successfully.")
    return redirect('student_management:subject_list')


# ─────────────────────────────────────────────
# QUESTIONS
# ─────────────────────────────────────────────

@admin_login_required
def question_list(request):
    questions = Question.objects.select_related('subject', 'submodule').order_by('-created_at')
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    submodules = SubModule.objects.filter(is_active=True).select_related('subject').order_by('subject__name', 'order', 'name')
    sources = Question.SOURCE_CHOICES
    years = Question.objects.exclude(year__isnull=True).values_list('year', flat=True).distinct().order_by('-year')

    subject_id = request.GET.get('subject')
    submodule_id = request.GET.get('submodule')
    source = request.GET.get('source')
    year = request.GET.get('year')

    if subject_id:
        questions = questions.filter(subject_id=subject_id)
    if submodule_id:
        questions = questions.filter(submodule_id=submodule_id)
    if source:
        questions = questions.filter(source=source)
    if year:
        questions = questions.filter(year=year)

    context = {
        'questions': questions,
        'subjects': subjects,
        'submodules':submodules,
        'sources': sources,
        'years': years,
        'selected_subject': subject_id,
        'selected_submodule': submodule_id,
        'selected_source': source,
        'selected_year': year,
    }
    return render(request, 'student_management/question_list.html', context)


@admin_login_required
def question_create(request):
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    submodules = SubModule.objects.none()
    library_assets = MediaLibrary.objects.filter(is_active=True).order_by('name')

    if request.method == 'POST':
        subject_id = request.POST.get('subject')
        submodule_id = request.POST.get('submodule', '').strip()
        question_text = request.POST.get('question_text', '').strip()
        option_a = request.POST.get('option_a', '').strip()
        option_b = request.POST.get('option_b', '').strip()
        option_c = request.POST.get('option_c', '').strip()
        option_d = request.POST.get('option_d', '').strip()
        option_e = request.POST.get('option_e', '').strip()
        correct_answer = request.POST.get('correct_answer', '').strip()
        source = request.POST.get('source', '').strip() or None
        year_raw = request.POST.get('year', '').strip()
        explanation = request.POST.get('explanation', '').strip()
        is_active = request.POST.get('is_active') == 'on'

        # Collect uploaded images (Option 1: direct upload)
        uploaded_images = {
            'QUESTION': request.FILES.get('question_image'),
            'OPTION_A': request.FILES.get('option_a_image'),
            'OPTION_B': request.FILES.get('option_b_image'),
            'OPTION_C': request.FILES.get('option_c_image'),
            'OPTION_D': request.FILES.get('option_d_image'),
            'OPTION_E': request.FILES.get('option_e_image'),
        }

        # Collect library asset IDs (Option 2: from library)
        library_ids = {
            'QUESTION': request.POST.get('question_library_id', '').strip(),
            'OPTION_A': request.POST.get('option_a_library_id', '').strip(),
            'OPTION_B': request.POST.get('option_b_library_id', '').strip(),
            'OPTION_C': request.POST.get('option_c_library_id', '').strip(),
            'OPTION_D': request.POST.get('option_d_library_id', '').strip(),
            'OPTION_E': request.POST.get('option_e_library_id', '').strip(),
        }

        errors = []
        subject = None
        submodule = None
        year = None

        if not subject_id:
            errors.append("Please select a subject.")
        else:
            subject = Subject.objects.filter(id=subject_id).first()
            if not subject:
                errors.append("Selected subject does not exist.")
        if submodule_id:
            submodule = SubModule.objects.filter(id=submodule_id).first()
            if not submodule:
                errors.append("Selected submodule does not exist.")
            elif subject and submodule.subject_id != subject.id:
                errors.append("Selected submodule does not belong to the chosen subject.")
        if not question_text:
            errors.append("Question text is required.")
        if not option_a:
            errors.append("Option A is required.")
        if not option_b:
            errors.append("Option B is required.")
        if not option_c:
            errors.append("Option C is required.")
        if not option_d:
            errors.append("Option D is required.")
        if not correct_answer:
            errors.append("Please select the correct answer.")

        if year_raw:
            try:
                year = int(year_raw)
                if year < 1900 or year > 2100:
                    errors.append("Please enter a valid year between 1900 and 2100.")
            except ValueError:
                errors.append("Year must be a valid number.")

        if source == 'PYQ' and not year:
            errors.append("Year is required when Source is PYQ.")

        if not errors and subject and question_text:
            normalized_current = normalize_question(question_text)
            comparison_source = source or 'OTHER'
            existing_qs = Question.objects.filter(subject=subject).values('question_text', 'source', 'year')
            is_duplicate = False
            for eq in existing_qs:
                if normalize_question(eq['question_text']) == normalized_current:
                    existing_source = eq['source'] or 'OTHER'
                    if comparison_source == 'PYQ' and existing_source == 'PYQ':
                        if year == eq['year']:
                            is_duplicate = True
                            break
                    elif comparison_source == existing_source and comparison_source != 'PYQ':
                        is_duplicate = True
                        break
            if is_duplicate:
                errors.append("This question already exists (duplicate detected based on subject, text, source, and year rules).")

        if errors:
            for err in errors:
                messages.error(request, err)
            return render(request, 'student_management/question_create.html', {
                'subjects': subjects,
                'submodules': submodules,
                'library_assets': library_assets,
                'post_data': request.POST,
            })

        # Create the question
        question = Question.objects.create(
            subject=subject,
            submodule=submodule,
            question_text=question_text,
            option_a=option_a,
            option_b=option_b,
            option_c=option_c,
            option_d=option_d,
            option_e=option_e or None,
            correct_answer=correct_answer,
            explanation=explanation or None,
            source=source,
            year=year,
            is_active=is_active,
        )

        FIELD_NAME_MAP = {
            'QUESTION': 'requires_question_image',
            'OPTION_A': 'requires_option_a_image',
            'OPTION_B': 'requires_option_b_image',
            'OPTION_C': 'requires_option_c_image',
            'OPTION_D': 'requires_option_d_image',
            'OPTION_E': 'requires_option_e_image',
        }

        media_created = False
        for media_type in ['QUESTION', 'OPTION_A', 'OPTION_B', 'OPTION_C', 'OPTION_D', 'OPTION_E']:
            file = uploaded_images.get(media_type)
            lib_id = library_ids.get(media_type)

            field_name = FIELD_NAME_MAP[media_type]

            if file:
                # Option 1: direct upload takes priority
                QuestionMedia.objects.create(
                    question=question,
                    media_type=media_type,
                    image=file,
                )
                setattr(question, field_name, True)
                media_created = True
            elif lib_id:
                # Option 2: from media library
                try:
                    asset = MediaLibrary.objects.get(id=lib_id, is_active=True)
                    QuestionMedia.objects.create(
                        question=question,
                        media_type=media_type,
                        media_library=asset,
                    )
                    setattr(question, field_name, True)
                    media_created = True
                except MediaLibrary.DoesNotExist:
                    pass

        if media_created:
            question.save(update_fields=[
                'requires_question_image', 'requires_option_a_image',
                'requires_option_b_image', 'requires_option_c_image',
                'requires_option_d_image', 'requires_option_e_image',
            ])
            _recalculate_media_status(question)

        messages.success(request, "Question created successfully.")
        return redirect('student_management:question_list')

    return render(request, 'student_management/question_create.html', {
        'subjects': subjects,
        'submodules': submodules,
        'library_assets': library_assets,
    })


@admin_login_required
def question_detail(request, id):
    question = get_object_or_404(Question.objects.select_related('subject'), id=id)
    media_map = {m.media_type: m for m in question.media_files.all()}

    question_options = [
        ('A', question.option_a),
        ('B', question.option_b),
        ('C', question.option_c),
        ('D', question.option_d),
    ]
    if question.option_e:
        question_options.append(('E', question.option_e))

    return render(request, 'student_management/question_detail.html', {
        'question': question,
        'question_options': question_options,
        'media_map': media_map,
    })


@admin_login_required
def question_edit(request, id):
    question = get_object_or_404(Question.objects.select_related('subject'), id=id)
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    library_assets = MediaLibrary.objects.filter(is_active=True).order_by('name')
    submodules = SubModule.objects.filter(subject_id=question.subject_id, is_active=True).order_by('order', 'name')

    if request.method == 'POST':
        subject_id = request.POST.get('subject')
        submodule_id = request.POST.get('submodule', '').strip()
        question_text = request.POST.get('question_text', '').strip()
        option_a = request.POST.get('option_a', '').strip()
        option_b = request.POST.get('option_b', '').strip()
        option_c = request.POST.get('option_c', '').strip()
        option_d = request.POST.get('option_d', '').strip()
        option_e = request.POST.get('option_e', '').strip()
        correct_answer = request.POST.get('correct_answer', '').strip()
        source = request.POST.get('source', '').strip() or None
        year_raw = request.POST.get('year', '').strip()
        explanation = request.POST.get('explanation', '').strip()
        is_active = request.POST.get('is_active') == 'on'

        errors = []
        subject = None
        submodule = None
        year = None

        if not subject_id:
            errors.append("Please select a subject.")
        else:
            subject = Subject.objects.filter(id=subject_id).first()
            if not subject:
                errors.append("Selected subject does not exist.")
        if submodule_id:
            submodule = SubModule.objects.filter(id=submodule_id).first()
            if not submodule:
                errors.append("Selected submodule does not exist.")
            elif subject and submodule.subject_id != subject.id:
                errors.append("Selected submodule does not belong to the chosen subject.")
        if not question_text:
            errors.append("Question text is required.")
        if not option_a:
            errors.append("Option A is required.")
        if not option_b:
            errors.append("Option B is required.")
        if not option_c:
            errors.append("Option C is required.")
        if not option_d:
            errors.append("Option D is required.")
        if not correct_answer:
            errors.append("Please select the correct answer.")

        if year_raw:
            try:
                year = int(year_raw)
                if year < 1900 or year > 2100:
                    errors.append("Please enter a valid year between 1900 and 2100.")
            except ValueError:
                errors.append("Year must be a valid number.")

        if source == 'PYQ' and not year:
            errors.append("Year is required when Source is PYQ.")

        if not errors and subject and question_text:
            normalized_current = normalize_question(question_text)
            comparison_source = source or 'OTHER'
            existing_qs = Question.objects.filter(subject=subject).exclude(id=id).values('question_text', 'source', 'year')
            is_duplicate = False
            for eq in existing_qs:
                if normalize_question(eq['question_text']) == normalized_current:
                    existing_source = eq['source'] or 'OTHER'
                    if comparison_source == 'PYQ' and existing_source == 'PYQ':
                        if year == eq['year']:
                            is_duplicate = True
                            break
                    elif comparison_source == existing_source and comparison_source != 'PYQ':
                        is_duplicate = True
                        break
            if is_duplicate:
                errors.append("This question already exists (duplicate detected based on subject, text, source, and year rules).")

        if errors:
            for err in errors:
                messages.error(request, err)
            media_map = {m.media_type: m for m in question.media_files.all()}
            return render(request, 'student_management/question_edit.html', {
                'question': question,
                'subjects': subjects,
                'media_map': media_map,
                'library_assets': library_assets,
                'submodules': submodules,
            })

        question.subject = subject
        question.submodule = submodule
        question.question_text = question_text
        question.option_a = option_a
        question.option_b = option_b
        question.option_c = option_c
        question.option_d = option_d
        question.option_e = option_e or None
        question.correct_answer = correct_answer
        question.explanation = explanation or None
        question.source = source
        question.year = year
        question.is_active = is_active
        question.save()

        # Process each media slot: clear or new upload or library pick
        media_map = {m.media_type: m for m in question.media_files.all()}
        requires_fields_to_save = []

        for field, media_type in SLOT_MAP:
            clear_key = f'clear_{media_type.lower()}_image'
            file_key = {
                'QUESTION': 'question_image',
                'OPTION_A': 'option_a_image',
                'OPTION_B': 'option_b_image',
                'OPTION_C': 'option_c_image',
                'OPTION_D': 'option_d_image',
                'OPTION_E': 'option_e_image',
            }[media_type]
            lib_id_key = {
                'QUESTION': 'question_library_id',
                'OPTION_A': 'option_a_library_id',
                'OPTION_B': 'option_b_library_id',
                'OPTION_C': 'option_c_library_id',
                'OPTION_D': 'option_d_library_id',
                'OPTION_E': 'option_e_library_id',
            }[media_type]

            if request.POST.get(clear_key) == 'on':
                # Delete the QuestionMedia record
                if media_type in media_map:
                    media_map[media_type].delete()
                setattr(question, field, False)
                requires_fields_to_save.append(field)
            elif request.FILES.get(file_key):
                # Direct file upload takes priority over library selection
                new_file = request.FILES[file_key]
                if media_type in media_map:
                    media_map[media_type].image = new_file
                    media_map[media_type].media_library = None
                    media_map[media_type].save()
                else:
                    QuestionMedia.objects.create(
                        question=question,
                        media_type=media_type,
                        image=new_file,
                    )
                setattr(question, field, True)
                requires_fields_to_save.append(field)
            else:
                lib_id = request.POST.get(lib_id_key, '').strip()
                if lib_id:
                    try:
                        asset = MediaLibrary.objects.get(id=lib_id, is_active=True)
                        if media_type in media_map:
                            media_map[media_type].image = ''
                            media_map[media_type].media_library = asset
                            media_map[media_type].save()
                        else:
                            QuestionMedia.objects.create(
                                question=question,
                                media_type=media_type,
                                media_library=asset,
                            )
                        setattr(question, field, True)
                        requires_fields_to_save.append(field)
                        # Clear any pending ref for this slot since it's now fulfilled
                        PendingMediaReference.objects.filter(
                            question=question, media_type=media_type
                        ).delete()
                    except MediaLibrary.DoesNotExist:
                        pass

        if requires_fields_to_save:
            question.save(update_fields=requires_fields_to_save)

        _recalculate_media_status(question)

        messages.success(request, "Question updated successfully.")
        return redirect('student_management:question_detail', id=question.id)

    media_map = {m.media_type: m for m in question.media_files.all()}
    return render(request, 'student_management/question_edit.html', {
        'question': question,
        'subjects': subjects,
        'submodules': submodules,
        'media_map': media_map,
        'library_assets': library_assets,
    })


@admin_login_required
def question_activate(request, id):
    question = get_object_or_404(Question, id=id)
    if question.is_active:
        messages.warning(request, "Question is already active.")
    else:
        question.is_active = True
        question.save()
        messages.success(request, "Question activated successfully.")
    return redirect('student_management:question_detail', id=question.id)


@admin_login_required
def question_deactivate(request, id):
    question = get_object_or_404(Question, id=id)
    if not question.is_active:
        messages.warning(request, "Question is already inactive.")
    else:
        question.is_active = False
        question.save()
        messages.warning(request, "Question deactivated successfully.")
    return redirect('student_management:question_detail', id=question.id)


@admin_login_required
def question_delete(request, id):
    question = get_object_or_404(Question, id=id)
    if request.method == 'POST':
        question.delete()
        messages.success(request, "Question deleted successfully.")
        return redirect('student_management:question_list')
    return redirect('student_management:question_detail', id=question.id)


@admin_login_required
def question_import(request):
    import_report = None

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            messages.error(request, "No file was uploaded. Please select an Excel file.")
            return redirect('student_management:question_import')

        filename = excel_file.name.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
            messages.error(request, "Invalid file format. Only .xlsx and .xls files are accepted.")
            return redirect('student_management:question_import')

        try:
            file_bytes = io.BytesIO(excel_file.read())
            df = pd.read_excel(file_bytes, engine='openpyxl')
        except Exception as e:
            messages.error(request, f"Could not read the Excel file. Please check the file is valid. ({e})")
            return redirect('student_management:question_import')

        required_columns = [
            'Subject', 'Question', 'Option A', 'Option B',
            'Option C', 'Option D', 'Correct Answer',
        ]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            messages.error(
                request,
                f"Missing required column(s): {', '.join(missing_columns)}. "
                "Please use the provided sample template."
            )
            return redirect('student_management:question_import')

        VALID_ANSWERS = {'A', 'B', 'C', 'D', 'E'}
        VALID_SOURCES = {'PYQ', 'EDUSETIN', 'OTHER'}

        # Optional media flag columns
        MEDIA_COLS = {
            'Has Question Image': 'requires_question_image',
            'Has Option A Image': 'requires_option_a_image',
            'Has Option B Image': 'requires_option_b_image',
            'Has Option C Image': 'requires_option_c_image',
            'Has Option D Image': 'requires_option_d_image',
            'Has Option E Image': 'requires_option_e_image',
        }

        # Optional image name columns (for MediaLibrary lookup)
        IMAGE_NAME_COLS = {
            'Question Image Name': 'QUESTION',
            'Option A Image Name': 'OPTION_A',
            'Option B Image Name': 'OPTION_B',
            'Option C Image Name': 'OPTION_C',
            'Option D Image Name': 'OPTION_D',
            'Option E Image Name': 'OPTION_E',
        }

        # Optional image URL columns (for MediaLibrary lookup)
        IMAGE_URL_COLS = {
            'Question Image URL': 'QUESTION',
            'Option A Image URL': 'OPTION_A',
            'Option B Image URL': 'OPTION_B',
            'Option C Image URL': 'OPTION_C',
            'Option D Image URL': 'OPTION_D',
            'Option E Image URL': 'OPTION_E',
        }

        FIELD_NAME_MAP = {
            'QUESTION': 'requires_question_image',
            'OPTION_A': 'requires_option_a_image',
            'OPTION_B': 'requires_option_b_image',
            'OPTION_C': 'requires_option_c_image',
            'OPTION_D': 'requires_option_d_image',
            'OPTION_E': 'requires_option_e_image',
        }

        imported_count = 0
        duplicate_count = 0
        failed_count = 0
        media_pending_count = 0
        error_rows = []
        duplicate_rows = []
        media_missing_rows = []  # detailed missing asset warnings

        existing_questions_cache = {}

        for row_num, (_, row) in enumerate(df.iterrows(), start=2):

            def get_cell(col_name, _row=row):
                val = _row.get(col_name, None)
                try:
                    if pd.isna(val):
                        return ''
                except (TypeError, ValueError):
                    pass
                return str(val).strip() if val is not None else ''

            subject_name   = get_cell('Subject')
            submodule_name = get_cell('Submodule') if 'Submodule' in df.columns else ''
            question_text  = get_cell('Question')
            option_a       = get_cell('Option A')
            option_b       = get_cell('Option B')
            option_c       = get_cell('Option C')
            option_d       = get_cell('Option D')
            option_e       = get_cell('Option E') if 'Option E' in df.columns else ''
            correct_answer = get_cell('Correct Answer').upper()
            source_raw     = get_cell('Source').upper() if 'Source' in df.columns else ''
            year_raw       = get_cell('Year') if 'Year' in df.columns else ''
            explanation    = get_cell('Explanation') if 'Explanation' in df.columns else ''

            # Parse media flag columns
            media_flags = {}
            for col, field in MEDIA_COLS.items():
                if col in df.columns:
                    media_flags[field] = _parse_bool_cell(row.get(col))
                else:
                    media_flags[field] = False

            # Parse image name columns
            image_names = {}
            for col, media_type in IMAGE_NAME_COLS.items():
                if col in df.columns:
                    val = get_cell(col, row)
                    image_names[media_type] = val if val else ''
                else:
                    image_names[media_type] = ''

            # Parse image URL columns
            image_urls = {}
            for col, media_type in IMAGE_URL_COLS.items():
                if col in df.columns:
                    val = get_cell(col, row)
                    image_urls[media_type] = val if val else ''
                else:
                    image_urls[media_type] = ''

            row_has_error = False
            subject = None

            def add_error(column, value, message):
                nonlocal row_has_error
                row_has_error = True
                error_rows.append({
                    'row': row_num,
                    'column': column,
                    'value': value if value else 'Empty',
                    'message': message,
                })

            if not subject_name:
                add_error('Subject', subject_name, 'Subject is required')
            else:
                subject = Subject.objects.filter(name__iexact=subject_name).first()
                if not subject:
                    add_error('Subject', subject_name, "Subject not found in database")
            submodule = None
            if submodule_name and subject:
                submodule = SubModule.objects.filter(subject=subject, name__iexact=submodule_name).first()
                if not submodule:
                    add_error('Submodule', submodule_name, "Submodule not found for the selected subject")
            if not question_text:
                add_error('Question', question_text, 'Question text is required')
            if not option_a:
                add_error('Option A', option_a, 'Option A is required')
            if not option_b:
                add_error('Option B', option_b, 'Option B is required')
            if not option_c:
                add_error('Option C', option_c, 'Option C is required')
            if not option_d:
                add_error('Option D', option_d, 'Option D is required')

            if not correct_answer:
                add_error('Correct Answer', correct_answer, 'Correct Answer is required')
            elif correct_answer not in VALID_ANSWERS:
                add_error('Correct Answer', correct_answer, 'Must be A, B, C, D, or E')

            source = None
            if source_raw:
                if source_raw not in VALID_SOURCES:
                    add_error('Source', source_raw, 'Must be PYQ, EDUSETIN, or OTHER')
                else:
                    source = source_raw

            year = None
            if year_raw:
                try:
                    year = int(float(year_raw))
                    if year < 1900 or year > 2100:
                        add_error('Year', year_raw, 'Must be between 1900 and 2100')
                        year = None
                except (ValueError, TypeError):
                    add_error('Year', year_raw, 'Must be a valid number')

            if source == 'PYQ' and not year:
                add_error('Year', year_raw, 'Year is required when Source is PYQ.')

            if row_has_error:
                failed_count += 1
                continue

            # Duplicate check
            if subject.id not in existing_questions_cache:
                existing_qs = Question.objects.filter(subject=subject).values('question_text', 'source', 'year')
                existing_questions_cache[subject.id] = [
                    {
                        'text': normalize_question(q['question_text']),
                        'source': q['source'] or 'OTHER',
                        'year': q['year']
                    }
                    for q in existing_qs
                ]

            normalized_current = normalize_question(question_text)
            comparison_source = source or 'OTHER'

            is_duplicate = False
            for eq in existing_questions_cache[subject.id]:
                if eq['text'] == normalized_current:
                    existing_source = eq['source']
                    if comparison_source == 'PYQ' and existing_source == 'PYQ':
                        if year == eq['year']:
                            is_duplicate = True
                            break
                    elif comparison_source == existing_source and comparison_source != 'PYQ':
                        is_duplicate = True
                        break

            if is_duplicate:
                duplicate_count += 1
                duplicate_rows.append({'row': row_num, 'message': 'Question already exists'})
                continue

            # Merge media_flags with image_name/url lookup:
            # If an image name or url column is provided, that slot also requires media.
            for media_type, img_name in image_names.items():
                if img_name:
                    field = FIELD_NAME_MAP[media_type]
                    media_flags[field] = True
            for media_type, img_url in image_urls.items():
                if img_url:
                    field = FIELD_NAME_MAP[media_type]
                    media_flags[field] = True

            # Create question with media flags
            try:
                has_any_media = any(media_flags.values())
                question = Question.objects.create(
                    subject=subject,
                    submodule=submodule,
                    question_text=question_text,
                    option_a=option_a,
                    option_b=option_b,
                    option_c=option_c,
                    option_d=option_d,
                    option_e=option_e or None,
                    correct_answer=correct_answer,
                    source=source,
                    year=year,
                    explanation=explanation or None,
                    is_active=True,
                    requires_question_image=media_flags.get('requires_question_image', False),
                    requires_option_a_image=media_flags.get('requires_option_a_image', False),
                    requires_option_b_image=media_flags.get('requires_option_b_image', False),
                    requires_option_c_image=media_flags.get('requires_option_c_image', False),
                    requires_option_d_image=media_flags.get('requires_option_d_image', False),
                    requires_option_e_image=media_flags.get('requires_option_e_image', False),
                    media_uploaded=False,
                )
                imported_count += 1

                # Process image lookups
                row_has_missing_media = False
                for media_type_tuple in SLOT_MAP:
                    media_type_enum = media_type_tuple[1]
                    img_name = image_names.get(media_type_enum, '')
                    img_url = image_urls.get(media_type_enum, '')
                    
                    if not img_name and not img_url:
                        continue
                        
                    asset = None
                    if img_name:
                        # Case-insensitive exact match
                        asset = MediaLibrary.objects.filter(name__iexact=img_name, is_active=True).first()
                    
                    if not asset and img_url:
                        # Django renames files on upload (e.g. brain.png → brain_XyZ.png),
                        # so matching against the physical image path stored in the DB will
                        # never work for user-supplied URLs. Instead, extract the filename
                        # from the URL and match against MediaLibrary.name (case-insensitive),
                        # which is exactly what the "Image Name" column already does.
                        #
                        # Supported input formats:
                        #   /media/media_library/brain.png
                        #   media_library/brain.png
                        #   http://127.0.0.1:8000/media/media_library/brain.png
                        import os as _os
                        from urllib.parse import urlparse as _urlparse
                        _parsed = _urlparse(img_url)
                        # Use the path component so full http(s):// URLs work too
                        _path = _parsed.path if _parsed.scheme else img_url
                        _filename = _os.path.basename(_path.rstrip('/'))
                        if _filename:
                            asset = MediaLibrary.objects.filter(name__iexact=_filename, is_active=True).first()

                    if asset:
                        # Create QuestionMedia referencing the library asset
                        QuestionMedia.objects.create(
                            question=question,
                            media_type=media_type_enum,
                            media_library=asset,
                        )
                        # Clear any pending ref if it existed
                        PendingMediaReference.objects.filter(
                            question=question, media_type=media_type_enum
                        ).delete()
                    else:
                        # Detailed warning for import report
                        if img_name:
                            col_label = {
                                'QUESTION': 'Question Image Name',
                                'OPTION_A': 'Option A Image Name',
                                'OPTION_B': 'Option B Image Name',
                                'OPTION_C': 'Option C Image Name',
                                'OPTION_D': 'Option D Image Name',
                                'OPTION_E': 'Option E Image Name',
                            }[media_type_enum]
                            val_to_report = img_name
                        else:
                            col_label = {
                                'QUESTION': 'Question Image URL',
                                'OPTION_A': 'Option A Image URL',
                                'OPTION_B': 'Option B Image URL',
                                'OPTION_C': 'Option C Image URL',
                                'OPTION_D': 'Option D Image URL',
                                'OPTION_E': 'Option E Image URL',
                            }[media_type_enum]
                            val_to_report = img_url
                            
                        # Asset not found — create PendingMediaReference
                        PendingMediaReference.objects.get_or_create(
                            question=question,
                            media_type=media_type_enum,
                            defaults={'expected_media_name': f"[URL] {val_to_report}" if not img_name else val_to_report},
                        )
                        media_missing_rows.append({
                            'row': row_num,
                            'column': col_label,
                            'value': val_to_report,
                            'message': 'Media asset not found in Media Library',
                        })
                        row_has_missing_media = True

                # Recalculate media_uploaded based on what was actually created
                _recalculate_media_status(question)

                if has_any_media:
                    # Check whether there are still pending slots
                    uploaded_types = set(question.media_files.values_list('media_type', flat=True))
                    required_types = {
                        mt for f, mt in SLOT_MAP if getattr(question, f)
                    }
                    if required_types - uploaded_types:
                        media_pending_count += 1

                existing_questions_cache[subject.id].append({
                    'text': normalized_current,
                    'source': source or 'OTHER',
                    'year': year
                })
            except Exception as e:
                failed_count += 1
                error_rows.append({
                    'row': row_num,
                    'column': '—',
                    'value': '—',
                    'message': f"Database error: {e}",
                })

        import_report = {
            'total_rows': len(df),
            'imported': imported_count,
            'duplicates': duplicate_count,
            'failed': failed_count,
            'media_pending': media_pending_count,
            'errors': error_rows,
            'duplicate_rows': duplicate_rows,
            'media_missing_rows': media_missing_rows,
        }

    return render(request, 'student_management/question_import.html', {
        'import_report': import_report,
    })


@admin_login_required
def download_question_template(request):
    """Generate and serve the Excel import template with all columns."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Questions"

        headers = [
            'Subject',
            'Submodule',
            'Question',
            'Option A',
            'Option B',
            'Option C',
            'Option D',
            'Option E',
            'Correct Answer',
            'Source',
            'Year',
            'Explanation',
            'Has Question Image',
            'Has Option A Image',
            'Has Option B Image',
            'Has Option C Image',
            'Has Option D Image',
            'Has Option E Image',
            'Question Image Name',
            'Option A Image Name',
            'Option B Image Name',
            'Option C Image Name',
            'Option D Image Name',
            'Option E Image Name',
            'Question Image URL',
            'Option A Image URL',
            'Option B Image URL',
            'Option C Image URL',
            'Option D Image URL',
            'Option E Image URL',
        ]

        header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
        new_col_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        new_cols = {
            'Question Image Name', 'Option A Image Name', 'Option B Image Name',
            'Option C Image Name', 'Option D Image Name', 'Option E Image Name',
            'Question Image URL', 'Option A Image URL', 'Option B Image URL',
            'Option C Image URL', 'Option D Image URL', 'Option E Image URL',
        }

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = center_align
            if header in new_cols:
                cell.fill = new_col_fill
            else:
                cell.fill = header_fill

        # Sample row
        sample_row = [
            'Mathematics',
            'Algebra',
            'What is 2 + 2?',
            '3',
            '4',
            '5',
            '6',
            '',
            'B',
            'EDUSETIN',
            '',
            'Basic addition.',
            'NO',
            'NO',
            'NO',
            'NO',
            'NO',
            'NO',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
        ]
        ws.append(sample_row)

        # Column widths
        col_widths = [20, 40, 20, 20, 20, 20, 20, 15, 12, 8, 30, 18, 18, 18, 18, 18, 18, 22, 22, 22, 22, 22, 22, 35, 35, 35, 35, 35, 35]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 35

        # Save to static directory
        file_dir = os.path.join(
            settings.BASE_DIR, 'student_management', 'static', 'student_management', 'downloads'
        )
        os.makedirs(file_dir, exist_ok=True)
        file_path = os.path.join(file_dir, 'sample_question_import.xlsx')
        wb.save(file_path)

        response = FileResponse(
            open(file_path, 'rb'),
            as_attachment=True,
            filename="sample_question_import.xlsx"
        )
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

    except Exception as e:
        messages.error(request, f"Could not generate template: {e}")
        return redirect('student_management:question_import')


# ─────────────────────────────────────────────
# MEDIA MANAGEMENT
# ─────────────────────────────────────────────

@admin_login_required
def media_pending(request):
    """List all questions that have media requirements but haven't been fully uploaded."""
    questions = Question.objects.select_related('subject').filter(
        Q(requires_question_image=True) |
        Q(requires_option_a_image=True) |
        Q(requires_option_b_image=True) |
        Q(requires_option_c_image=True) |
        Q(requires_option_d_image=True) |
        Q(requires_option_e_image=True),
        media_uploaded=False,
    ).prefetch_related('media_files', 'pending_media_refs').order_by('id')

    pending_data = []
    for question in questions:
        media_map = {m.media_type: m for m in question.media_files.all()}
        slots = _get_media_slot_status(question, media_map)
        required_slots = [s for s in slots if s['required']]
        uploaded_slots = [s for s in required_slots if s['uploaded']]
        pending_data.append({
            'question': question,
            'slots': required_slots,
            'required_count': len(required_slots),
            'uploaded_count': len(uploaded_slots),
        })

    return render(request, 'student_management/media_pending.html', {
        'pending_data': pending_data,
        'total_pending': len(pending_data),
    })


@admin_login_required
def media_upload(request, id):
    """Upload / manage media for a single question."""
    question = get_object_or_404(Question.objects.select_related('subject'), id=id)

    if request.method == 'POST':
        media_map = {m.media_type: m for m in question.media_files.all()}
        requires_fields_to_save = []

        for field, media_type in SLOT_MAP:
            clear_key = f'clear_{media_type.lower()}'
            file_key = {
                'QUESTION': 'question_image',
                'OPTION_A': 'option_a_image',
                'OPTION_B': 'option_b_image',
                'OPTION_C': 'option_c_image',
                'OPTION_D': 'option_d_image',
                'OPTION_E': 'option_e_image',
            }[media_type]
            lib_id_key = {
                'QUESTION': 'question_library_id',
                'OPTION_A': 'option_a_library_id',
                'OPTION_B': 'option_b_library_id',
                'OPTION_C': 'option_c_library_id',
                'OPTION_D': 'option_d_library_id',
                'OPTION_E': 'option_e_library_id',
            }[media_type]

            if request.POST.get(clear_key) == 'on':
                if media_type in media_map:
                    media_map[media_type].delete()
                setattr(question, field, False)
                requires_fields_to_save.append(field)
            elif request.FILES.get(file_key):
                new_file = request.FILES[file_key]
                if media_type in media_map:
                    media_map[media_type].image = new_file
                    media_map[media_type].media_library = None
                    media_map[media_type].save()
                else:
                    QuestionMedia.objects.create(
                        question=question,
                        media_type=media_type,
                        image=new_file,
                    )
                setattr(question, field, True)
                requires_fields_to_save.append(field)
                # Clear pending ref if slot fulfilled
                PendingMediaReference.objects.filter(
                    question=question, media_type=media_type
                ).delete()
            else:
                lib_id = request.POST.get(lib_id_key, '').strip()
                if lib_id:
                    try:
                        asset = MediaLibrary.objects.get(id=lib_id, is_active=True)
                        if media_type in media_map:
                            media_map[media_type].image = ''
                            media_map[media_type].media_library = asset
                            media_map[media_type].save()
                        else:
                            QuestionMedia.objects.create(
                                question=question,
                                media_type=media_type,
                                media_library=asset,
                            )
                        setattr(question, field, True)
                        requires_fields_to_save.append(field)
                        # Clear pending ref
                        PendingMediaReference.objects.filter(
                            question=question, media_type=media_type
                        ).delete()
                    except MediaLibrary.DoesNotExist:
                        pass

        if requires_fields_to_save:
            question.save(update_fields=requires_fields_to_save)

        _recalculate_media_status(question)
        messages.success(request, f"Media for Question #{question.id} saved successfully.")

        action = request.POST.get('action', 'save')
        if action == 'save_next':
            next_question = Question.objects.filter(
                Q(requires_question_image=True) |
                Q(requires_option_a_image=True) |
                Q(requires_option_b_image=True) |
                Q(requires_option_c_image=True) |
                Q(requires_option_d_image=True) |
                Q(requires_option_e_image=True),
                media_uploaded=False,
            ).order_by('id').first()

            if next_question:
                return redirect('student_management:media_upload', id=next_question.id)
            else:
                messages.success(request, "All pending media uploads completed.")
                return redirect('student_management:media_pending')

        return redirect('student_management:media_upload', id=question.id)

    # GET
    media_map = {m.media_type: m for m in question.media_files.all()}
    slots = _get_media_slot_status(question, media_map)
    required_count = sum(1 for s in slots if s['required'])
    uploaded_count = sum(1 for s in slots if s['required'] and s['uploaded'])
    library_assets = MediaLibrary.objects.filter(is_active=True).order_by('name')

    # Find next pending question for Save & Next button context
    next_question = Question.objects.filter(
        Q(requires_question_image=True) |
        Q(requires_option_a_image=True) |
        Q(requires_option_b_image=True) |
        Q(requires_option_c_image=True) |
        Q(requires_option_d_image=True) |
        Q(requires_option_e_image=True),
        media_uploaded=False,
    ).exclude(id=id).order_by('id').first()

    return render(request, 'student_management/media_upload.html', {
        'question': question,
        'slots': slots,
        'media_map': media_map,
        'required_count': required_count,
        'uploaded_count': uploaded_count,
        'next_question': next_question,
        'library_assets': library_assets,
    })


@admin_login_required
def media_delete(request, id, media_type):
    """Delete a single QuestionMedia record."""
    question = get_object_or_404(Question, id=id)

    if request.method == 'POST':
        media_type = media_type.upper()
        try:
            qm = QuestionMedia.objects.get(question=question, media_type=media_type)
            qm.delete()

            # Update the requires_* flag
            field_map = {
                'QUESTION': 'requires_question_image',
                'OPTION_A': 'requires_option_a_image',
                'OPTION_B': 'requires_option_b_image',
                'OPTION_C': 'requires_option_c_image',
                'OPTION_D': 'requires_option_d_image',
                'OPTION_E': 'requires_option_e_image',
            }
            field = field_map.get(media_type)
            if field:
                setattr(question, field, False)
                question.save(update_fields=[field])

            _recalculate_media_status(question)
            messages.success(request, f"{SLOT_LABELS.get(media_type, media_type)} deleted.")
        except QuestionMedia.DoesNotExist:
            messages.warning(request, "Media record not found.")

    return redirect('student_management:media_upload', id=id)


@admin_login_required
def media_library(request):
    """
    Refactored Media Library page with two sections:
    1. Media Assets Repository (MediaLibrary model)
    2. Question Media Usage (QuestionMedia model)
    """
    # Section 1: MediaLibrary assets
    assets_qs = MediaLibrary.objects.all().order_by('name')

    # Section 2: QuestionMedia usage records
    usage_qs = QuestionMedia.objects.select_related(
        'question__subject', 'media_library'
    ).order_by('-created_at')

    subjects = Subject.objects.filter(is_active=True).order_by('name')
    media_type_choices = QuestionMedia.MEDIA_TYPES

    # Filters for usage section
    subject_id = request.GET.get('subject')
    media_type_filter = request.GET.get('media_type')
    search = request.GET.get('search', '').strip()
    asset_search = request.GET.get('asset_search', '').strip()
    active_filter = request.GET.get('active_filter', '')

    # Asset filters
    if asset_search:
        assets_qs = assets_qs.filter(
            Q(name__icontains=asset_search) |
            Q(description__icontains=asset_search)
        )
    if active_filter == '1':
        assets_qs = assets_qs.filter(is_active=True)
    elif active_filter == '0':
        assets_qs = assets_qs.filter(is_active=False)

    # Usage filters
    if subject_id:
        usage_qs = usage_qs.filter(question__subject_id=subject_id)
    if media_type_filter:
        usage_qs = usage_qs.filter(media_type=media_type_filter)
    if search:
        usage_qs = usage_qs.filter(question__question_text__icontains=search)

    active_tab = request.GET.get('tab', 'assets')

    return render(request, 'student_management/media_library.html', {
        # Assets section
        'assets_list': assets_qs,
        'assets_count': assets_qs.count(),
        'asset_search': asset_search,
        'active_filter': active_filter,
        # Usage section
        'media_list': usage_qs,
        'total_count': usage_qs.count(),
        'subjects': subjects,
        'media_type_choices': media_type_choices,
        'selected_subject': subject_id,
        'selected_media_type': media_type_filter,
        'search': search,
        'active_tab': active_tab,
    })


# ─────────────────────────────────────────────
# MEDIA ASSETS CRUD
# ─────────────────────────────────────────────

@admin_login_required
def media_assets_upload(request):
    """Upload a new MediaLibrary asset."""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        image_file = request.FILES.get('image')

        errors = []

        if not name:
            errors.append("Asset name is required.")

        if not image_file:
            errors.append("Image file is required.")
        else:
            if not _validate_image_extension(image_file.name):
                errors.append("Invalid image format. Allowed formats: jpg, jpeg, png, webp.")

        # Case-insensitive duplicate name check
        if name and MediaLibrary.objects.filter(name__iexact=name).exists():
            errors.append("A media asset with this name already exists.")

        if errors:
            for err in errors:
                messages.error(request, err)
            return render(request, 'student_management/media_assets_upload.html', {
                'post_data': request.POST,
            })

        MediaLibrary.objects.create(
            name=name,
            image=image_file,
            description=description or None,
            is_active=is_active,
        )
        messages.success(request, f"Media asset \"{name}\" uploaded successfully.")
        return redirect('student_management:media_library')

    return render(request, 'student_management/media_assets_upload.html', {})


@admin_login_required
def media_assets_edit(request, id):
    """Edit a MediaLibrary asset's name, description, active status."""
    asset = get_object_or_404(MediaLibrary, id=id)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        image_file = request.FILES.get('image')

        errors = []

        if not name:
            errors.append("Asset name is required.")

        # Case-insensitive duplicate name check (excluding self)
        if name and MediaLibrary.objects.filter(name__iexact=name).exclude(id=id).exists():
            errors.append("A media asset with this name already exists.")

        if image_file and not _validate_image_extension(image_file.name):
            errors.append("Invalid image format. Allowed formats: jpg, jpeg, png, webp.")

        if errors:
            for err in errors:
                messages.error(request, err)
            return render(request, 'student_management/media_assets_edit.html', {'asset': asset})

        asset.name = name
        asset.description = description or None
        asset.is_active = is_active
        if image_file:
            asset.image = image_file
        asset.save()
        messages.success(request, f"Media asset \"{name}\" updated successfully.")
        return redirect('student_management:media_library')

    return render(request, 'student_management/media_assets_edit.html', {'asset': asset})


@admin_login_required
def media_assets_delete(request, id):
    """Delete a MediaLibrary asset — blocked if currently in use."""
    asset = get_object_or_404(MediaLibrary, id=id)

    if request.method == 'POST':
        if asset.is_in_use:
            messages.error(
                request,
                "Cannot delete this media asset because it is currently used by one or more questions."
            )
            return redirect('student_management:media_library')

        name = asset.name
        asset.delete()
        messages.success(request, f"Media asset \"{name}\" deleted successfully.")
        return redirect('student_management:media_library')

    return redirect('student_management:media_library')


@admin_login_required
def media_assets_toggle(request, id):
    """Toggle active/inactive status of a MediaLibrary asset."""
    asset = get_object_or_404(MediaLibrary, id=id)
    if request.method == 'POST':
        asset.is_active = not asset.is_active
        asset.save()
        status = "activated" if asset.is_active else "deactivated"
        messages.success(request, f"Media asset \"{asset.name}\" {status}.")
    return redirect('student_management:media_library')

@admin_login_required
def media_assets_api(request):
    """
    JSON API for the library picker modal.
    Returns active MediaLibrary assets matching search query.
    Searches name and description (case-insensitive, partial match).
    """
    search = request.GET.get('q', '').strip()
    qs = MediaLibrary.objects.filter(is_active=True)
    if search:
        qs = qs.filter(
            Q(name__icontains=search) |
            Q(description__icontains=search)
        )
    qs = qs.order_by('name')[:50]

    results = []
    for asset in qs:
        results.append({
            'id': asset.id,
            'name': asset.name,
            'description': asset.description or '',
            'image_url': asset.image.url if asset.image else '',
            'usage_count': asset.usage_count,
        })
    return JsonResponse({'results': results})

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django import forms
from .models import ExamType, Exam

class ExamTypeForm(forms.ModelForm):
    class Meta:
        model = ExamType
        fields = ['name', 'description', 'is_active']
        widgets = {
            'name': forms.Select(choices=[
                ('', '-- Select Type --'),
                ('PYQ',    'PYQ Test'),
                ('MOCK',   'Mock Test'),
                ('QUIZ',   'Quick Quiz'),
                ('CUSTOM', 'Custom'),
            ]),
        }

class ExamTypeForm(forms.ModelForm):
    name = forms.ChoiceField(
        choices=[('', '-- Select Type --')] + ExamType.TYPE_CHOICES,
        widget=forms.Select(attrs={'class': 'form-select'}),
    )

    class Meta:
        model = ExamType
        fields = ['name', 'description', 'is_active']

class ExamForm(forms.ModelForm):
    class Meta:
        model = Exam
        fields = [
            'exam_type', 'title', 'description',
            'duration_minutes', 'marks_per_question', 'negative_marks',
            'total_questions', 'pyq_years',
            'subjects', 'submodules', 'selected_questions',
            'is_active', 'image',
        ]
        widgets = {
            'subjects': forms.CheckboxSelectMultiple(),
            'submodules': forms.CheckboxSelectMultiple(),
            'selected_questions': forms.CheckboxSelectMultiple(),
            'description': forms.Textarea(attrs={'rows': 3}),
            'image': forms.ClearableFileInput(attrs={'class': 'form-control', 'accept': 'image/*'}),
        }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['subjects'].queryset = Subject.objects.filter(is_active=True)
        self.fields['submodules'].queryset = SubModule.objects.filter(
            is_active=True
        ).select_related('subject').order_by('subject__name', 'order', 'name')
        self.fields['selected_questions'].queryset = Question.objects.filter(is_active=True)
        self.fields['description'].required = False
        self.fields['submodules'].required = False

@admin_login_required
def examtype_list(request):
    exam_types = ExamType.objects.all()
    add_form = ExamTypeForm()
    edit_forms = {et.pk: ExamTypeForm(instance=et) for et in exam_types}

    if request.method == 'POST':
        action = request.POST.get('action')
        pk = request.POST.get('pk')

        if action == 'add':
            form = ExamTypeForm(request.POST)
            if form.is_valid():
                instance = form.save(commit=False)
                if instance.name == 'CUSTOM':
                    instance.custom_name = request.POST.get('custom_name', '').strip() or None
                else:
                    instance.custom_name = None
                instance.save()
                messages.success(request, 'Exam type added.')
                return redirect('student_management:examtype_list')
            else:
                add_form = form

        elif action == 'edit' and pk:
            instance = get_object_or_404(ExamType, pk=pk)
            form = ExamTypeForm(request.POST, instance=instance)
            if form.is_valid():
                obj = form.save(commit=False)
                if obj.name == 'CUSTOM':
                    obj.custom_name = request.POST.get('custom_name', '').strip() or None
                else:
                    obj.custom_name = None
                obj.save()
                messages.success(request, 'Exam type updated.')
                return redirect('student_management:examtype_list')
            else:
                edit_forms[instance.pk] = form

        elif action == 'delete' and pk:
            instance = get_object_or_404(ExamType, pk=pk)
            instance.delete()
            messages.success(request, 'Exam type deleted.')
            return redirect('student_management:examtype_list')

    edit_forms_list = [(et.pk, edit_forms[et.pk]) for et in exam_types]
    return render(request, 'student_management/examtype_list.html', {
        'exam_types': exam_types,
        'add_form': add_form,
        'edit_forms_list': edit_forms_list,
    })

@admin_login_required
def exam_list(request):
    exams = Exam.objects.select_related('exam_type').prefetch_related('subjects', 'submodules')

    if request.method == 'POST' and request.POST.get('action') == 'delete':
        pk = request.POST.get('pk')
        exam = get_object_or_404(Exam, pk=pk)
        exam.delete()
        messages.success(request, 'Exam deleted.')
        return redirect('student_management:exam_list')

    return render(request, 'student_management/exam_list.html', {'exams': exams})
@admin_login_required
def exam_add(request):
    form = ExamForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Exam created.')
        return redirect('student_management:exam_list')  # ← fixed
    return render(request, 'student_management/exam_add.html', {'form': form})


@admin_login_required
def exam_edit(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    form = ExamForm(request.POST or None, request.FILES or None, instance=exam)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Exam updated.')
        return redirect('student_management:exam_list')  # ← fixed
    return render(request, 'student_management/exam_edit.html', {'form': form, 'exam': exam})



import json
from django.http import JsonResponse
 

@admin_login_required
def ajax_subjects_by_examtype(request):
    """
    GET /student_management/ajax/subjects/?exam_type_id=<id>
    Returns all active subjects (no ExamType→Subject relation in current model).
    """
    if request.headers.get("X-Requested-With") != "XMLHttpRequest":
        return JsonResponse({"error": "Bad request"}, status=400)
 
    exam_type_id = request.GET.get("exam_type_id")
    if not exam_type_id:
        return JsonResponse({"subjects": []})
 
    try:
        ExamType.objects.get(pk=exam_type_id)
    except ExamType.DoesNotExist:
        return JsonResponse({"subjects": []})
 
    # No ExamType→Subject link in your model, so return all active subjects.
    # If you add a M2M later, swap this line for:
    #   subjects = exam_type.subjects.filter(is_active=True)
    subjects = Subject.objects.filter(is_active=True).order_by("name")
    data = [{"id": s.pk, "name": str(s)} for s in subjects]
    return JsonResponse({"subjects": data})
 
 
@admin_login_required
def ajax_questions_by_filter(request):
    # Define questions FIRST before any if blocks
    questions = Question.objects.filter(is_active=True).select_related('subject', 'submodule')

    years_only = request.GET.get('years_only')
    if years_only:
        years = list(
            questions.exclude(year__isnull=True)
            .values_list('year', flat=True)
            .distinct()
            .order_by('-year')
        )
        return JsonResponse({'years': years})

    subject_ids = request.GET.getlist('subject_ids')
    submodule_ids = request.GET.getlist('submodule_ids')
    year_filters = request.GET.getlist('years')
    exclude_used = request.GET.get('exclude_used')
    exam_id = request.GET.get('exam_id')

    if subject_ids:
        questions = questions.filter(subject_id__in=subject_ids)
    if submodule_ids:
        questions = questions.filter(submodule_id__in=submodule_ids)
    if year_filters:
        questions = questions.filter(year__in=year_filters)

    if exclude_used:
        other_exams = Exam.objects.all()
        if exam_id:
            other_exams = other_exams.exclude(pk=exam_id)
        questions = questions.exclude(exams__in=other_exams).distinct()

    data = [
        {
            'id':        q.id,
            'text':      q.question_text,
            'subject':   q.subject.name if q.subject else None,
            'submodule': q.submodule.name if q.submodule else None,
            'year':      q.year,
            'source':    q.source,
        }
        for q in questions.order_by('subject__name', 'id')
    ]
    return JsonResponse({'questions': data})

@admin_login_required
def submodule_list(request):
    empty_slug_submodules = SubModule.objects.filter(slug='')
    for sm in empty_slug_submodules:
        sm.save()

    submodules = SubModule.objects.select_related('subject').all()

    search = request.GET.get('search', '').strip()
    subject_id = request.GET.get('subject', '').strip()
    status = request.GET.get('status', '').strip()

    if search:
        submodules = submodules.filter(
            Q(name__icontains=search) | Q(subject__name__icontains=search)
        )

    if subject_id:
        submodules = submodules.filter(subject_id=subject_id)

    if status == 'active':
        submodules = submodules.filter(is_active=True)
    elif status == 'inactive':
        submodules = submodules.filter(is_active=False)

    paginator = Paginator(submodules, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'submodules': page_obj.object_list,
        'subjects': Subject.objects.filter(is_active=True).order_by('name'),
        'search': search,
        'selected_subject': subject_id,
        'selected_status': status,
    }
    return render(request, 'student_management/submodule_list.html', context)

@admin_login_required
def submodule_add(request):
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    initial = {
        'subject_id': '',
        'name': '',
        'description': '',
        'order': 0,
        'is_active': True,
    }

    if request.method == 'POST':
        subject_id = request.POST.get('subject')
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        order = request.POST.get('order') or 0
        is_active = request.POST.get('is_active') == 'on'
        image = request.FILES.get('image')

        errors = {}

        if not subject_id:
            errors['subject'] = 'Subject is required.'
        if not name:
            errors['name'] = 'Name is required.'
        elif SubModule.objects.filter(subject_id=subject_id, name__iexact=name).exists():
            errors['name'] = 'A submodule with this name already exists for the selected subject.'

        try:
            order = int(order)
        except (TypeError, ValueError):
            errors['order'] = 'Order must be a number.'

        if not errors:
            submodule = SubModule.objects.create(
                subject_id=subject_id,
                name=name,
                description=description,
                order=order,
                is_active=is_active,
                image=image,
            )
            messages.success(request, f'Submodule "{submodule.name}" was created successfully.')
            return redirect('student_management:submodule_list')
        else:
            for field, error in errors.items():
                messages.error(request, error)

            initial = {
                'subject_id': subject_id or '',
                'name': name,
                'description': description,
                'order': order,
                'is_active': is_active,
            }
            context = {
                'subjects': subjects,
                'submodule': None,
                'initial': initial,
                'errors': errors,
            }
            return render(request, 'student_management/submodule_form.html', context)

    context = {'subjects': subjects, 'submodule': None, 'initial': initial, 'errors': {}}
    return render(request, 'student_management/submodule_form.html', context)



@admin_login_required
def submodule_edit(request, slug):
    submodule = get_object_or_404(SubModule, slug=slug)
    subjects = Subject.objects.filter(is_active=True).order_by('name')
    initial = {
        'subject_id': submodule.subject_id,
        'name': submodule.name,
        'description': submodule.description or '',
        'order': submodule.order,
        'is_active': submodule.is_active,
    }

    if request.method == 'POST':
        subject_id = request.POST.get('subject')
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        order = request.POST.get('order') or 0
        is_active = request.POST.get('is_active') == 'on'
        image = request.FILES.get('image')
        remove_image = request.POST.get('remove_image') == 'on'

        errors = {}

        if not subject_id:
            errors['subject'] = 'Subject is required.'
        if not name:
            errors['name'] = 'Name is required.'
        elif SubModule.objects.filter(subject_id=subject_id, name__iexact=name).exclude(pk=submodule.pk).exists():
            errors['name'] = 'A submodule with this name already exists for the selected subject.'

        try:
            order = int(order)
        except (TypeError, ValueError):
            errors['order'] = 'Order must be a number.'

        if not errors:
            submodule.subject_id = subject_id
            submodule.name = name
            submodule.description = description
            submodule.order = order
            submodule.is_active = is_active

            if remove_image and submodule.image:
                submodule.image.delete(save=False)
                submodule.image = None

            if image:
                if submodule.image:
                    submodule.image.delete(save=False)
                submodule.image = image

            submodule.save()
            messages.success(request, f'Submodule "{submodule.name}" was updated successfully.')
            return redirect('student_management:submodule_list')
        else:
            for field, error in errors.items():
                messages.error(request, error)

            initial = {
                'subject_id': subject_id or '',
                'name': name,
                'description': description,
                'order': order,
                'is_active': is_active,
            }
            context = {
                'subjects': subjects,
                'submodule': submodule,
                'initial': initial,
                'errors': errors,
            }
            return render(request, 'student_management/submodule_form.html', context)

    context = {'subjects': subjects, 'submodule': submodule, 'initial': initial, 'errors': {}}
    return render(request, 'student_management/submodule_form.html', context)



@admin_login_required
def submodule_delete(request, slug):
    submodule = get_object_or_404(SubModule, slug=slug)

    if request.method == 'POST':
        name = submodule.name
        if submodule.image:
            submodule.image.delete(save=False)
        submodule.delete()
        messages.success(request, f'Submodule "{name}" was deleted successfully.')
    else:
        messages.error(request, 'Invalid request method.')

    return redirect('student_management:submodule_list')
@login_required(login_url='admin_login')
def submodules_by_subject_api(request):
    subject_id = request.GET.get('subject_id', '').strip()
    if not subject_id:
        return JsonResponse({'submodules': []})
    submodules = SubModule.objects.filter(
        subject_id=subject_id,
        is_active=True
    ).order_by('name').values('id', 'name')
    return JsonResponse({'submodules': list(submodules)})


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q
from django import forms
from .models import SubscriptionPlan, Subject, SubModule, Exam


class SubscriptionPlanForm(forms.ModelForm):
    subjects = forms.ModelMultipleChoiceField(
        queryset=Subject.objects.filter(is_active=True),
        widget=forms.CheckboxSelectMultiple,
        required=False,
    )
    submodules = forms.ModelMultipleChoiceField(
        queryset=SubModule.objects.filter(is_active=True),
        widget=forms.CheckboxSelectMultiple,
        required=False,
    )
    exams = forms.ModelMultipleChoiceField(
        queryset=Exam.objects.filter(is_active=True),
        widget=forms.CheckboxSelectMultiple,
        required=False,
    )

    class Meta:
        model = SubscriptionPlan
        fields = ["name", "price", "duration_days", "subjects", "submodules", "exams", "is_active"]


@admin_login_required
def plan_list(request):
    query = request.GET.get("q", "")
    plans = SubscriptionPlan.objects.prefetch_related("subjects", "exams", "submodules")
    if query:
        plans = plans.filter(Q(name__icontains=query) | Q(description__icontains=query))
    return render(request, "student_management/plan_list.html", {"plans": plans, "query": query})


@admin_login_required
def plan_detail(request, pk):
    plan = get_object_or_404(SubscriptionPlan, pk=pk)
    return render(request, "student_management/plan_detail.html", {"plan": plan})


@admin_login_required
def plan_create(request):
    form = SubscriptionPlanForm(request.POST or None)
    if form.is_valid():
        plan = form.save()
        messages.success(request, f'Plan "{plan.name}" created successfully.')
        return redirect("student_management:plan_list")
    return render(request, "student_management/plan_form.html", {"form": form, "title": "Create Plan"})


@admin_login_required
def plan_update(request, pk):
    plan = get_object_or_404(SubscriptionPlan, pk=pk)
    form = SubscriptionPlanForm(request.POST or None, instance=plan)
    if form.is_valid():
        plan = form.save()
        messages.success(request, f'Plan "{plan.name}" updated successfully.')
        return redirect("student_management:plan_list")
    return render(request, "student_management/plan_form.html", {"form": form, "title": "Edit Plan", "plan": plan})


@admin_login_required
def plan_delete(request, pk):
    plan = get_object_or_404(SubscriptionPlan, pk=pk)
    name = plan.name
    plan.delete()
    messages.success(request, f'Plan "{name}" deleted.')
    return redirect("student_management:plan_list")

from student_portal.models import Student
from .models import Payment, SubscriptionPlan


@admin_login_required
def payment_list(request):
    payments = Payment.objects.select_related('student', 'plan').order_by('-created_at')

    search    = request.GET.get('search', '').strip()
    status    = request.GET.get('status', '').strip()
    plan_id   = request.GET.get('plan', '').strip()

    if search:
        payments = payments.filter(
            Q(student__full_name__icontains=search) |
            Q(plan__name__icontains=search) |
            Q(transaction_id__icontains=search)
        )
    if status:
        payments = payments.filter(status=status)
    if plan_id:
        payments = payments.filter(plan_id=plan_id)

    plans = SubscriptionPlan.objects.filter(is_active=True).order_by('name')

    return render(request, 'student_management/payment_list.html', {
        'payments': payments,
        'plans': plans,
        'search': search,
        'selected_status': status,
        'selected_plan': plan_id,
        'status_choices': Payment.STATUS_CHOICES,
    })


@admin_login_required
def payment_detail(request, id):
    payment = get_object_or_404(
        Payment.objects.select_related('student', 'plan'), id=id
    )
    return render(request, 'student_management/payment_detail.html', {
        'payment': payment,
    })


@admin_login_required
def payment_create(request):
    students = Student.objects.filter(is_active=True).order_by('full_name')
    plans    = SubscriptionPlan.objects.filter(is_active=True).order_by('name')

    if request.method == 'POST':
        student_id     = request.POST.get('student', '').strip()
        plan_id        = request.POST.get('plan', '').strip()
        amount         = request.POST.get('amount', '').strip()
        transaction_id = request.POST.get('transaction_id', '').strip()
        status         = request.POST.get('status', Payment.STATUS_PENDING).strip()

        errors  = []
        student = None
        plan    = None

        if not student_id:
            errors.append("Please select a student.")
        else:
            student = Student.objects.filter(id=student_id).first()
            if not student:
                errors.append("Selected student does not exist.")

        if not plan_id:
            errors.append("Please select a plan.")
        else:
            plan = SubscriptionPlan.objects.filter(id=plan_id).first()
            if not plan:
                errors.append("Selected plan does not exist.")

        if not amount:
            errors.append("Amount is required.")
        else:
            try:
                amount = float(amount)
                if amount < 0:
                    errors.append("Amount cannot be negative.")
            except ValueError:
                errors.append("Amount must be a valid number.")

        if errors:
            for err in errors:
                messages.error(request, err)
            return render(request, 'student_management/payment_form.html', {
                'students': students,
                'plans': plans,
                'status_choices': Payment.STATUS_CHOICES,
                'post_data': request.POST,
                'title': 'Create Payment',
            })

        from django.utils import timezone
        from datetime import timedelta

        payment = Payment.objects.create(
            student=student,
            plan=plan,
            amount=amount,
            transaction_id=transaction_id,
            status=status,
            paid_at=timezone.now() if status == Payment.STATUS_SUCCESS else None,
            expires_at=(
                timezone.now() + timedelta(days=plan.duration_days)
                if status == Payment.STATUS_SUCCESS else None
            ),
        )
        messages.success(request, f"Payment for {student.full_name} created successfully.")
        return redirect('student_management:payment_detail', id=payment.id)

    return render(request, 'student_management/payment_form.html', {
        'students': students,
        'plans': plans,
        'status_choices': Payment.STATUS_CHOICES,
        'title': 'Create Payment',
    })


@admin_login_required
def payment_update_status(request, id):
    payment = get_object_or_404(Payment, id=id)

    if request.method == 'POST':
        new_status     = request.POST.get('status', '').strip()
        valid_statuses = [s[0] for s in Payment.STATUS_CHOICES]

        if new_status not in valid_statuses:
            messages.error(request, "Invalid status selected.")
            return redirect('student_management:payment_detail', id=payment.id)

        from django.utils import timezone
        from datetime import timedelta

        payment.status = new_status

        if new_status == Payment.STATUS_SUCCESS and not payment.paid_at:
            payment.paid_at    = timezone.now()
            payment.expires_at = timezone.now() + timedelta(days=payment.plan.duration_days)

        payment.save()
        messages.success(request, f"Payment status updated to {payment.get_status_display()}.")

    return redirect('student_management:payment_detail', id=payment.id)


@admin_login_required
def payment_delete(request, id):
    payment = get_object_or_404(Payment, id=id)
    if request.method == 'POST':
        payment.delete()
        messages.success(request, "Payment record deleted successfully.")
        return redirect('student_management:payment_list')
    return redirect('student_management:payment_detail', id=payment.id)